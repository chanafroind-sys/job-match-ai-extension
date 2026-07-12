import asyncio
import hashlib
import json
import os
import re
import time
import urllib.parse
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Optional

import httpx
from anthropic import AsyncAnthropic
from json_repair import repair_json
from dotenv import load_dotenv
from fastapi import FastAPI, Header, HTTPException, Request, WebSocket, WebSocketDisconnect
from fastapi.responses import RedirectResponse, StreamingResponse
from starlette.types import ASGIApp, Receive, Scope, Send
from pydantic import BaseModel

load_dotenv()

BACKEND_URL = os.getenv("BACKEND_URL", "https://job-match-ai-extension.onrender.com")
ANTHROPIC_API_KEY: str = os.environ["ANTHROPIC_API_KEY"]

# Gumroad — product permalink used as product_id in the API (both work)
GUMROAD_PRODUCT_PERMALINK: str = os.getenv("GUMROAD_PRODUCT_PERMALINK", "job-match-ai")
GUMROAD_ACCESS_TOKEN: str   = os.getenv("GUMROAD_ACCESS_TOKEN", "")   # seller token, for future use
GUMROAD_SELLER_HANDLE: str  = os.getenv("GUMROAD_SELLER_HANDLE", "expert-dev-ai")
# Premium is a variant of the same product — append ?variant=Premium so Gumroad
# pre-selects the Premium tier on the checkout page.
UPGRADE_URL: str = (
    f"https://{GUMROAD_SELLER_HANDLE}.gumroad.com/l/{GUMROAD_PRODUCT_PERMALINK}"
    "?variant=Premium"
)

# Keep old env var name for backward compat with existing Render config
GUMROAD_PRODUCT_ID: str = os.getenv("GUMROAD_PRODUCT_ID", GUMROAD_PRODUCT_PERMALINK)

MAX_DEVICES_PER_KEY: int = int(os.getenv("MAX_DEVICES_PER_KEY", "3"))
MONTHLY_USAGE_LIMIT: int = int(os.getenv("MONTHLY_USAGE_LIMIT", "100"))
USAGE_FILE = Path(__file__).parent / "usage.json"
CLICKS_FILE = Path(__file__).parent / "clicks.json"

# ── WebSocket push-notification registry (in-memory, reset on restart) ────────
# user_id (sha256[:16] of license key) → set of active WebSocket connections
_ws_connections: dict[str, set] = {}
# app_id → {"user_id": str, "job_title": str, "company": str}
_app_id_registry: dict[str, dict] = {}

def _ws_user_id(license_key: str) -> str:
    """Stable, short, opaque identifier for a license key."""
    return hashlib.sha256(license_key.encode()).hexdigest()[:16]

async def _ws_push(user_id: str, payload: dict) -> None:
    """Send payload to all active WebSocket connections for this user."""
    conns = _ws_connections.get(user_id, set())
    if not conns:
        return
    msg = json.dumps(payload)
    dead: set = set()
    for ws in list(conns):
        try:
            await ws.send_text(msg)
        except Exception:
            dead.add(ws)
    conns -= dead
    if not conns:
        _ws_connections.pop(user_id, None)

# ── Premium ───────────────────────────────────────────────────────────────────

# Comma-separated list of license keys that get premium access
# regardless of Gumroad variants (useful for admin/dev keys)
_PREMIUM_KEYS_ENV = os.getenv("PREMIUM_KEYS", "")
STATIC_PREMIUM_KEYS: set = {k.strip() for k in _PREMIUM_KEYS_ENV.split(",") if k.strip()}

RAW_JOBS_FILE = Path(__file__).parent / "raw_jobs.json"
MAX_RAW_JOBS = 10_000


def _load_raw_jobs() -> list:
    try:
        return json.loads(RAW_JOBS_FILE.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def _save_raw_jobs(jobs: list) -> None:
    RAW_JOBS_FILE.write_text(json.dumps(jobs, ensure_ascii=False), encoding="utf-8")


def _quick_filter(jobs: list, cv_text: str) -> list:
    """Fast keyword pre-filter — no AI, removes obviously irrelevant jobs."""
    STOP = {
        "with","that","this","have","from","they","will","been","their","what","when",
        "more","also","into","than","then","some","which","them","these","those","would",
        "there","were","your","like","just","over","such","each","after","about","using",
        "work","years","team","skills","looking","good","strong","plus","great","know",
        "high","well","make","help","need","able","both","must","very","only","much",
    }
    cv_words = {w.lower() for w in re.findall(r'\b[a-zA-Z]{4,}\b', cv_text)} - STOP
    if len(cv_words) < 5:
        return jobs  # too few CV keywords — pass everything through
    filtered = []
    for job in jobs:
        text_lower = (job.get("text") or "").lower()
        if sum(1 for w in cv_words if w in text_lower) >= 2:
            filtered.append(job)
    return filtered

# ── Link tracking ─────────────────────────────────────────────────────────────

# Single-pass regex — three alternatives, evaluated in order:
#   group(1)+group(2) : markdown link  [display text](https://url)
#   group(3)          : raw URL with protocol  https://... or http://...
#   group(4)          : bare domain/path without protocol  github.com/user, mysite.io/portfolio, etc.
_LINK_RE = re.compile(
    r'\[([^\]]+)\]\((https?://[^\)]+)\)'                                          # markdown link
    r'|(https?://[^\s\)\]\"\'>]+)'                                                # raw URL w/ protocol
    r'|(?<![/\w@.])'                                                              # bare URL — not preceded by URL chars
      r'([a-zA-Z0-9][a-zA-Z0-9\-]{2,}\.(?:com|io|dev|me|net|org|co|ai|app|tech)'
      r'/[^\s\)\]\"\'>]+)'
)

def inject_tracking_links(cv_text: str, app_id: str) -> str:
    """Replace all URLs with tracking links.

    Handles three forms in order:
      1. [LINK:display|url]  — already-formatted tokens (Claude preserved from original CV);
                               rewrap the inner URL with a tracking URL, keep the display text.
      2. [text](url)         — markdown link
      3. https://...         — raw URL with protocol
      4. domain.com/path     — bare domain URL without protocol
    """
    def _make(url: str, display: str) -> str:
        url = re.sub(r'[.,;:!?]+$', '', url)
        display = re.sub(r'[.,;:!?]+$', '', display)
        full_url = url if url.startswith("http") else f"https://{url}"
        target = ("github" if "github.com" in full_url
                  else "linkedin" if "linkedin.com" in full_url
                  else "portfolio")
        enc = urllib.parse.quote(full_url, safe="")
        tracking = f"{BACKEND_URL}/api/v1/track?app_id={app_id}&target={target}&url={enc}"
        return f"[LINK:{display}|{tracking}]"

    # Pass 1 — rewrap any [LINK:display|url] tokens that Claude preserved verbatim.
    # Results are stashed behind NUL-delimited placeholders so that Pass 2's regex
    # cannot see (and double-wrap) the tracking URLs embedded inside them.
    _EXISTING_RE = re.compile(r'\[LINK:([^\|]*)\|([^\]]+)\]')
    _stash: dict[str, str] = {}

    def _pass1(m: re.Match) -> str:
        key = f"\x00{len(_stash)}\x00"
        _stash[key] = _make(m.group(2), m.group(1) or m.group(2))
        return key

    cv_text = _EXISTING_RE.sub(_pass1, cv_text)

    # Pass 2 — wrap any remaining raw / bare / markdown URLs Claude introduced.
    def _replace(m: re.Match) -> str:
        if m.group(1):                          # markdown [text](url)
            return _make(m.group(2), m.group(1))
        if m.group(3):                          # raw URL with protocol
            url = m.group(3)
            display = re.sub(r'^https?://', '', url).rstrip('/')
            return _make(url, display)
        url = m.group(4)                        # bare domain/path
        return _make(url, url.rstrip('/'))

    cv_text = _LINK_RE.sub(_replace, cv_text)

    # Restore Pass 1 results
    for key, val in _stash.items():
        cv_text = cv_text.replace(key, val)

    return cv_text

anthropic_client = AsyncAnthropic(api_key=ANTHROPIC_API_KEY)


# ── Usage tracking ────────────────────────────────────────────────────────────

def _load_usage() -> dict:
    try:
        return json.loads(USAGE_FILE.read_text())
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_usage(data: dict) -> None:
    USAGE_FILE.write_text(json.dumps(data, indent=2))


def _month_key() -> str:
    now = datetime.utcnow()
    return f"{now.year}-{now.month:02d}"


def get_usage_count(license_key: str) -> int:
    usage = _load_usage()
    return usage.get(license_key, {}).get(_month_key(), 0)


def increment_usage(license_key: str) -> int:
    usage = _load_usage()
    month = _month_key()
    usage.setdefault(license_key, {})[month] = usage.get(license_key, {}).get(month, 0) + 1
    _save_usage(usage)
    return usage[license_key][month]


# ── License cache (in-memory, TTL-based) ──────────────────────────────────────
# Keys are stored as SHA-256 hashes — never plaintext license strings in memory.

_license_cache: dict[str, dict] = {}
_CACHE_TTL_VALID   = 3600   # 1 h  — valid licenses
_CACHE_TTL_INVALID = 30     # 30 s — failed lookups (short so mis-typed or mis-matched keys retry quickly)

def _ck(license_key: str) -> str:
    return hashlib.sha256(license_key.strip().lower().encode()).hexdigest()

def _cache_get(license_key: str) -> dict | None:
    """Return cached entry or None if absent / expired."""
    entry = _license_cache.get(_ck(license_key))
    if not entry:
        return None
    ttl = _CACHE_TTL_VALID if entry["valid"] else _CACHE_TTL_INVALID
    if time.monotonic() - entry["ts"] > ttl:
        _license_cache.pop(_ck(license_key), None)
        return None
    return entry

def _cache_set(license_key: str, data: dict | None, *, valid: bool) -> None:
    _license_cache[_ck(license_key)] = {"data": data, "ts": time.monotonic(), "valid": valid}


# ── Gumroad verification ──────────────────────────────────────────────────────

TEST_LICENSE_KEY = "TEST-MICHAL-FAKE-KEY"


async def verify_gumroad_license(license_key: str) -> dict:
    """
    Verify license via Gumroad API.

    Returns a dict with:
      email, uses, tier ("standard"|"premium"), isPremium (bool), subscriptionActive (bool)

    Results are cached: 1 h for valid keys, 5 min for invalid ones.
    On Gumroad network errors, serves stale cache so existing users aren't locked out.
    """
    masked = license_key[:4] + "****" if len(license_key) > 4 else "****"
    print(f"[JMA:verify] key={masked} len={len(license_key.strip())}")

    # ── Static / admin key bypass (TEST key + PREMIUM_KEYS env var) ──────────
    k = license_key.strip()
    if k == TEST_LICENSE_KEY:
        print("[JMA:verify] TEST KEY — bypass OK")
        return {"email": "test@internal", "uses": 1, "tier": "premium",
                "isPremium": True, "subscriptionActive": True}
    if k in STATIC_PREMIUM_KEYS:
        print(f"[JMA:verify] STATIC PREMIUM KEY — bypass OK key={masked}")
        return {"email": "admin@internal", "uses": 1, "tier": "premium",
                "isPremium": True, "subscriptionActive": True}

    # ── Cache hit ─────────────────────────────────────────────────────────────
    cached = _cache_get(license_key)
    if cached is not None:
        if not cached["valid"]:
            print(f"[JMA:verify] CACHED invalid key={masked}")
            raise HTTPException(status_code=403, detail="Invalid or expired license key.")
        print(f"[JMA:verify] CACHE HIT key={masked} tier={cached['data'].get('tier')}")
        return cached["data"]

    # ── Gumroad API call — try all known product IDs ──────────────────────────
    # A license key is tied to whichever product the user purchased from.
    # We try GUMROAD_PRODUCT_PERMALINK first, then fall back to GUMROAD_PRODUCT_ID
    # (which may hold an older permalink like "oechku") so users who bought before
    # a product rename are not locked out.
    _product_ids_to_try: list[str] = list(dict.fromkeys(filter(None, [
        GUMROAD_PRODUCT_PERMALINK,   # current permalink  (e.g. "job-match-ai")
        GUMROAD_PRODUCT_ID,          # legacy env-var value (e.g. "oechku")
    ])))

    data: dict = {}
    last_network_error: Exception | None = None

    for pid in _product_ids_to_try:
        try:
            # Gumroad v2 API accepts either product_permalink or product_id
            # (they're the same value — the URL slug). Try both field names
            # to maximise compatibility across API versions.
            base_payload: dict = {
                "license_key": license_key.strip(),
                "increment_uses_count": "false",
            }
            if GUMROAD_ACCESS_TOKEN:
                base_payload["access_token"] = GUMROAD_ACCESS_TOKEN

            headers: dict = {}
            if GUMROAD_ACCESS_TOKEN:
                headers["Authorization"] = f"Bearer {GUMROAD_ACCESS_TOKEN}"

            for field in ("product_permalink", "product_id"):
                payload = {field: pid, **base_payload}
                async with httpx.AsyncClient(timeout=12) as client:
                    resp = await client.post(
                        "https://api.gumroad.com/v2/licenses/verify",
                        data=payload,
                        headers=headers,
                    )
                data = resp.json()
                print(f"[JMA:verify] Gumroad field={field!r} pid={pid!r} "
                      f"status={resp.status_code} success={data.get('success')} "
                      f"msg={data.get('message','')!r}")
                if data.get("success"):
                    break
            if data.get("success"):
                break   # found the right product+field combo — stop outer loop
        except Exception as e:
            print(f"[JMA:verify] Gumroad network error pid={pid!r}: {e}")
            last_network_error = e

    if not data.get("success"):
        if last_network_error and not data:
            # All attempts were network failures — try stale cache as grace fallback
            stale = _license_cache.get(_ck(license_key))
            if stale and stale["valid"] and stale["data"]:
                print(f"[JMA:verify] serving stale cache (Gumroad unreachable) key={masked}")
                return stale["data"]
            raise HTTPException(status_code=503,
                detail="לא הצלחנו להגיע לשירות אימות הרישיון. נסי שוב בעוד רגע.")

        print(f"[JMA:verify] all product_ids rejected key={masked} tried={_product_ids_to_try}")
        _cache_set(license_key, None, valid=False)
        raise HTTPException(status_code=403, detail="Invalid or expired license key.")

    purchase: dict = data.get("purchase") or {}

    # ── Subscription lifecycle checks ─────────────────────────────────────────
    sub_failed   = purchase.get("subscription_failed_at")
    chargebacked = purchase.get("chargebacked", False)
    refunded     = purchase.get("refunded", False)
    sub_cancelled = purchase.get("subscription_cancelled_at")

    if sub_failed:
        print(f"[JMA:verify] payment failed: {sub_failed}")
        _cache_set(license_key, None, valid=False)
        raise HTTPException(status_code=403,
            detail="Subscription payment failed. Please update your payment method on Gumroad.")

    if chargebacked or refunded:
        print(f"[JMA:verify] chargeback={chargebacked} refunded={refunded}")
        _cache_set(license_key, None, valid=False)
        raise HTTPException(status_code=403,
            detail="This license is no longer valid (refunded or disputed).")

    # Cancelled but billing period not yet over — Gumroad will return success=false
    # once it expires. We flag it so the UI can show a "cancels soon" warning.
    subscription_active = not bool(sub_cancelled)

    # ── Device / activation limit ─────────────────────────────────────────────
    uses: int = purchase.get("uses", 0)
    if uses > MAX_DEVICES_PER_KEY:
        print(f"[JMA:verify] too many devices: uses={uses} max={MAX_DEVICES_PER_KEY}")
        raise HTTPException(status_code=403,
            detail=f"This license key is active on {uses} devices (max {MAX_DEVICES_PER_KEY}). Please purchase a separate license.")

    # ── Tier detection ────────────────────────────────────────────────────────
    # Gumroad returns variants as a JSON-encoded string: '{"Plan": "Premium"}'
    variants_raw = purchase.get("variants", "")
    try:
        variants = json.loads(variants_raw) if isinstance(variants_raw, str) and variants_raw else (variants_raw or {})
    except (json.JSONDecodeError, TypeError):
        variants = {}

    plan_raw = (variants.get("Plan") or variants.get("plan") or "").strip().lower()
    tier     = "premium" if plan_raw == "premium" else "standard"
    is_premium = tier == "premium" or license_key.strip() in STATIC_PREMIUM_KEYS

    print(f"[JMA:verify] OK email={purchase.get('email','')} uses={uses} "
          f"tier={tier!r} subscriptionActive={subscription_active}")

    result = {
        "email":              purchase.get("email", ""),
        "uses":               uses,
        "tier":               tier,
        "isPremium":          is_premium,
        "subscriptionActive": subscription_active,
    }
    _cache_set(license_key, result, valid=True)
    return result


async def _verify_premium(license_key: str) -> bool:
    """Premium gate check — uses cache, admin keys bypass Gumroad instantly."""
    k = license_key.strip()
    if k == TEST_LICENSE_KEY or k in STATIC_PREMIUM_KEYS:
        return True
    try:
        info = await verify_gumroad_license(k)
        return info.get("isPremium", False)
    except HTTPException:
        raise
    except Exception:
        return False


async def require_license(license_key: str) -> str:
    """Validate license and enforce monthly usage cap. Returns the license key on success."""
    if not license_key or not license_key.strip():
        print("[JMA:license] REJECTED — empty key")
        raise HTTPException(status_code=401,
            detail="No license key provided. Please enter a valid license key in the extension settings.")
    await verify_gumroad_license(license_key)
    count = get_usage_count(license_key)
    print(f"[JMA:license] usage={count}/{MONTHLY_USAGE_LIMIT}")
    if count >= MONTHLY_USAGE_LIMIT:
        print("[JMA:license] REJECTED — monthly limit reached")
        raise HTTPException(status_code=429,
            detail=f"Monthly usage limit reached ({MONTHLY_USAGE_LIMIT} analyses). Resets on the 1st of next month.")
    return license_key


# ── Claude API ────────────────────────────────────────────────────────────────

# async def call_claude(prompt: str, max_tokens: int = 1200, model: str = "claude-haiku-4-5-20251001") -> str:
#     print(f"[JMA:claude] calling model prompt_len={len(prompt)} max_tokens={max_tokens}")
#     last_exc: Exception | None = None
#     for attempt in range(1, 4):
#         try:
#             message = await anthropic_client.messages.create(
#                 model="claude-haiku-4-5-20251001",
#                 max_tokens=max_tokens,
#                 messages=[{"role": "user", "content": prompt}],
#             )
#             result = "".join(block.text for block in message.content if hasattr(block, "text"))
#             print(f"[JMA:claude] attempt={attempt} response_len={len(result)} stop_reason={message.stop_reason}")
#             return result
#         except Exception as e:
#             last_exc = e
#             print(f"[JMA:claude] attempt={attempt} ERROR: {type(e).__name__}: {e}")
#             if attempt < 3:
#                 await asyncio.sleep(3)
#     raise last_exc
async def call_claude(prompt: str, max_tokens: int = 1200, model: str = "claude-haiku-4-5-20251001",
                       check_truncation: bool = False) -> str:
    print(f"[JMA:claude] calling model prompt_len={len(prompt)} max_tokens={max_tokens}")
    last_exc: Exception | None = None
    for attempt in range(1, 4):
        try:
            message = await anthropic_client.messages.create(
                model=model,
                max_tokens=max_tokens,
                messages=[{"role": "user", "content": prompt}],
            )
            result = "".join(block.text for block in message.content if hasattr(block, "text"))
            print(f"[JMA:claude] attempt={attempt} response_len={len(result)} stop_reason={message.stop_reason}")
            if check_truncation and message.stop_reason == "max_tokens":
                raise HTTPException(status_code=422,
                    detail="התוכן שנוצר ארוך מדי וקוצץ באמצע. אנא נסי שוב או קצרי את קורות החיים המקוריים.")
            return result
        except HTTPException:
            raise
        except Exception as e:
            last_exc = e
            print(f"[JMA:claude] attempt={attempt} ERROR: {type(e).__name__}: {e}")
            if attempt < 3:
                await asyncio.sleep(3)
    raise last_exc

async def call_claude_cached(
    system_blocks: list,
    user_content: str,
    max_tokens: int = 1200,
    model: str = "claude-sonnet-4-6",
) -> tuple[str, str]:
    """Call Claude with a cached system prompt. Returns (text, stop_reason)."""
    total_sys = sum(len(b.get("text", "")) for b in system_blocks)
    print(f"[JMA:claude] cached call sys_len={total_sys} user_len={len(user_content)} max_tokens={max_tokens} model={model}")
    last_exc: Exception | None = None
    for attempt in range(1, 4):
        try:
            message = await anthropic_client.messages.create(
                model=model,
                max_tokens=max_tokens,
                system=system_blocks,
                messages=[{"role": "user", "content": user_content}],
            )
            result = "".join(block.text for block in message.content if hasattr(block, "text"))
            stop   = message.stop_reason or "end_turn"
            usage  = message.usage
            print(
                f"[JMA:claude] attempt={attempt} response_len={len(result)} "
                f"stop={stop} "
                f"cache_read={getattr(usage,'cache_read_input_tokens',0)} "
                f"cache_write={getattr(usage,'cache_creation_input_tokens',0)}"
            )
            return result, stop
        except Exception as e:
            last_exc = e
            print(f"[JMA:claude] attempt={attempt} ERROR: {type(e).__name__}: {e}")
            if attempt < 3:
                await asyncio.sleep(3)
    raise last_exc


def _extract_raw(text: str, opening: str = "{", closing: str = "}") -> str:
    """Pull the first JSON object/array out of LLM text, stripping fences."""
    # Strip markdown code fences first
    fence = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
    if fence:
        return fence.group(1).strip()
    # Find the outermost {…} or […]
    start = text.find(opening)
    if start == -1:
        return text.strip()
    depth, end = 0, start
    close_ch = closing
    open_ch  = opening
    for i, ch in enumerate(text[start:], start):
        if ch == open_ch:
            depth += 1
        elif ch == close_ch:
            depth -= 1
            if depth == 0:
                end = i
                break
    return text[start:end + 1]


def _safe_json_loads(raw: str) -> any:
    """json.loads with json_repair fallback for LLM-produced JSON."""
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        repaired = repair_json(raw, return_objects=True)
        # repair_json returns the parsed object when return_objects=True
        return repaired


def parse_json_response(text: str) -> dict:
    raw = _extract_raw(text, "{", "}")
    result = _safe_json_loads(raw)
    if not isinstance(result, dict):
        print(f"[JMA:parse] expected dict, got {type(result)} | raw_start={raw[:200]!r}")
        raise ValueError("LLM did not return a JSON object")
    return result


def parse_json_array(text: str) -> list:
    """Parse Claude's response as a JSON array, stripping any markdown fences."""
    raw = _extract_raw(text, "[", "]")
    result = _safe_json_loads(raw)
    return result if isinstance(result, list) else []


def parse_cv_sections_py(cv_text: str) -> dict:
    """Parse a structured CV (with [SECTION] markers) into {marker: content}."""
    markers = {'[NAME]', '[HEADLINE]', '[CONTACT]', '[PROFILE]',
               '[EXPERIENCE]', '[EDUCATION]', '[SKILLS]', '[LANGUAGES]'}
    sections: dict = {}
    current: str | None = None
    lines: list = []
    for line in cv_text.splitlines():
        clean = line.strip().lstrip('#').strip().strip('*').strip()
        if clean in markers:
            if current:
                sections[current] = '\n'.join(lines).strip()
            current, lines = clean, []
        elif current:
            lines.append(line)
    if current:
        sections[current] = '\n'.join(lines).strip()
    return sections

def _validate_cv_output(cv_final: str, original_cv: str) -> list[str]:
    """Cheap heuristic checks — does NOT call the AI. Flags likely content loss."""
    issues: list[str] = []
    sections = parse_cv_sections_py(cv_final)

    for marker in ['[NAME]', '[PROFILE]', '[EXPERIENCE]', '[EDUCATION]', '[SKILLS]']:
        if not sections.get(marker, '').strip():
            issues.append(f"missing_or_empty:{marker}")

    exp_final = sections.get('[EXPERIENCE]', '')
    if exp_final.strip() and exp_final.count('$ ') == 0:
        issues.append("experience_has_no_bullets")

    orig_len = len(original_cv.strip())
    final_len = len(cv_final.strip())
    if orig_len > 300 and final_len < orig_len * 0.25:
        issues.append(f"drastic_shrink orig_chars={orig_len} final_chars={final_len}")

    return issues
# ── Prompts ───────────────────────────────────────────────────────────────────

# ── Preflight system block builder ───────────────────────────────────────────

def _cv_system_blocks(cv_text: str) -> list:
    """Return a system-prompt block list with the CV cached via ephemeral cache_control."""
    return [
        {
            "type": "text",
            "text": (
                "You are a senior recruitment expert and screening specialist.\n\n"
                "=== CANDIDATE CV ===\n"
                f"{cv_text}\n"
                "=== END CV ==="
            ),
            "cache_control": {"type": "ephemeral"},
        }
    ]


# ── Pass-1 prompt: fast base score + gap percentage ───────────────────────────

_SCORING_RULES = """\
SCORING RULES:
STEP 1 — classify every requirement:
  CRITICAL = "must"/"required"/"mandatory" or listed in the primary requirements section
  SECONDARY = "nice to have"/"preferred"/"advantage"/"bonus" or in a lower-priority section

STEP 2 — deduct from 100:
  Missing CRITICAL requirement: -15 to -25 per item (proportional to centrality)
  Partial/theoretical on a CRITICAL item: -5 to -12
  Years shortfall: (required - actual) / required × 30 pts
  Seniority mismatch (Senior/Lead in title but not in CV): cap at 65
  Missing SECONDARY/nice-to-have: -2 to -5 per item (NEVER more than -5 each)

DOMAIN MISMATCH: cap at 55 ONLY when the candidate genuinely lacks the critical skills.
  If they have the required skills but come from a different job-title background, do NOT cap — score on skill fit alone.

STEP 3 — positive partial offsets (secondary gaps only):
  Strong relevant academics +5 | Adjacent transferable skills +5 | Relevant projects +5

CALIBRATION: 85+ shortlist | 70-84 interview | 55-69 gaps | <40 wrong fit
A candidate meeting all CRITICALs but lacking all SECONDARYs → 65-75, not below 55.
Score <50 means most CRITICALs are missing — not just a different background.\
"""

BASE_ANALYSIS_USER = """Analyse the fit between the candidate CV (in your system prompt) and this job posting.

{scoring_rules}

Additionally compute:
  gap_pct = integer 0-40 — how many score points COULD be gained if the candidate could fully clarify all uncertain/missing areas through follow-up questions. 0 means the CV is fully self-explanatory; 35 means strong potential to improve with answers.

MANDATORY: Before setting base_score you MUST complete the scoring_reasoning field with explicit step-by-step arithmetic covering ALL four sections below. Do NOT skip or summarise — write the actual numbers.

Return ONLY valid JSON — no markdown:
{{
  "scoring_reasoning": "<REQUIRED — explicit maths in English covering all four steps:\n1. TOTAL YEARS CHECK: job requires X total yrs, candidate has Y → shortfall Z → penalty = Z/X×30 = P pts\n2. CORE TECH YEAR CHECK: for each tech with an explicit years requirement, write 'TechName: required Xyr, candidate has Yyr → shortfall Z → penalty P pts'\n3. MISSING CRITICAL TECHS: list every CRITICAL technology/tool not found anywhere in the CV and its individual deduction (−15 to −25). Write 'none' only if literally nothing is missing.\n4. FINAL ARITHMETIC: start=100, list every deduction with its label, sum them, clamp to [0,100] → base_score=N>",
  "base_score": <integer 0-100>,
  "gap_pct": <integer 0-40>,
  "jobTitle": "<job title>",
  "company": "<company name>",
  "jobLanguage": "hebrew" | "english",
  "summary": "<2 sentences in Hebrew describing overall fit>",
  "strengths": ["<Hebrew sentence with skill names in English>", ...],
  "hard_gaps": ["<Hebrew sentence with skill names in English>", ...]
}}

=== JOB DESCRIPTION ===
{job_text}"""


# ── Pass-2 prompt: weighted questions ─────────────────────────────────────────

QUESTIONS_USER = """The candidate CV is in your system prompt.
The initial match score is {base_score}% with {gap_pct} improvement points available through clarifying questions.

Identify the {n_questions} most important skills or experiences that are unclear or missing in the CV and are clearly required or strongly preferred by this job.

For EACH question assign an integer `weight` (1-{gap_pct}). The weights MUST sum to exactly {gap_pct}.

ALL text (question, explanation) MUST be in Hebrew. Skill names stay in English.

Return ONLY valid JSON — no markdown, no extra fields:
{{"questions":[
  {{"id":"q1","skill":"<English>","question":"<שאלה מקצועית קצרה>","explanation":"<הסבר ממוקד עד 15 מילים>","weight":<integer>}},
  ...
]}}

=== JOB DESCRIPTION ===
{job_text}"""


# kept for backward-compat with non-preflight full analysis path
QUESTIONS_PROMPT = ""  # unused — see BASE_ANALYSIS_USER / QUESTIONS_USER above

ANALYZE_PROMPT = """You are a senior recruitment expert. Analyze the fit between the candidate's CV and the job posting.

MANDATORY LANGUAGE RULE: The fields summary, strengths, and hard_gaps MUST be written in Hebrew (עברית) — no exceptions, regardless of the job's language. Technical terms, skill names, and technologies stay in English inside the Hebrew sentences.

SCORING RULES — read carefully before assigning a score:

STEP 1 — classify every requirement in the job posting:
  CRITICAL = explicitly labelled "must", "required", "mandatory", or listed in the primary requirements section
  SECONDARY = labelled "nice to have", "preferred", "advantage", "bonus", or listed in a separate lower-priority section

STEP 2 — deduct from 100:
  Missing CRITICAL requirement: -15 to -25 per item (how central it is to the role)
  Partial / theoretical match on a CRITICAL item: -5 to -12
  Years-of-experience shortfall: (required - actual) / required × 30 pts
  Seniority mismatch (Senior/Lead in title but not in CV): cap score at 65
  Missing SECONDARY / nice-to-have: -2 to -5 per item (NEVER more than -5 each)

DOMAIN MISMATCH — apply ONLY when the candidate lacks the critical skills themselves, not merely because their job title or background sounds different:
  - If the candidate's background is in a different domain BUT they demonstrably possess the CRITICAL skills listed in the job, do NOT apply any domain-mismatch cap. Score purely on skill fit.
  - Apply "cap at 55" ONLY when the candidate's skills are genuinely misaligned with the critical requirements (e.g. an embedded C developer applying for a React frontend role, with no transferable skills at all).
  - Example: a backend developer applying for a BI role who has SQL + Power BI + ETL knowledge is NOT a domain mismatch — she has the required skills. Her background in AI does not penalise her.

STEP 3 — positive partial offsets (apply only to secondary gaps):
  Strong relevant academic background: +5
  Adjacent skills that transfer: +5
  Relevant projects / portfolio: +5

CALIBRATION: 85+ strong shortlist | 70-84 interview-worthy | 55-69 real gaps | <40 wrong fit
A candidate who meets all CRITICAL requirements but lacks all SECONDARY ones should score 65-75, not below 55.
A score below 50 means the candidate is fundamentally missing most of the CRITICAL requirements — not merely that their background sounds different.

- Return JSON only (no markdown, no explanations):
{{
  "score": <0-100>,
  "jobTitle": "<job title from posting>",
  "company": "<company name>",
  "jobLanguage": "hebrew" | "english",
  "summary": "<2 sentences in Hebrew — MANDATORY Hebrew>",
  "strengths": ["<Hebrew sentence, skill names in English>", ...],
  "hard_gaps": ["<Hebrew sentence, skill name in English>", ...]
}}
{answers_section}
=== CV ===
{cv_text}
=== JOB DESCRIPTION ===
{job_text}"""

_CV_LANG_RULE_EN = """LANGUAGE RULE — 100% ENGLISH ONLY:
- The entire CV must be written in English from start to finish — no exceptions
- The candidate's name must always appear in English only — never translate it to Hebrew or any other language
- Do not mix any Hebrew words or characters anywhere in the document"""

_CV_LANG_RULE_HE = """LANGUAGE RULE — HEBREW OUTPUT:
- Write the entire CV in Hebrew (עברית) — this is mandatory
- The candidate's name: if it is a Hebrew name write it in Hebrew; if it is a Latin/English name keep it as-is
- Keep technical terms in English: programming languages (Python, Java, C++), frameworks (React, FastAPI, Node.js), tools (Docker, Git, AWS), company names, product names, software names
- Translate job titles and all section content to Hebrew (e.g. "Software Engineer" → "מהנדס תוכנה")
- Write in natural, professional Hebrew — not a word-for-word translation
- The document direction is right-to-left"""

_CV_LANG_CHECK_EN = "1. LANGUAGE: Is the entire CV in 100% English? If any Hebrew words or characters appear anywhere — translate or remove them. The candidate's name must remain in English exactly as written."

_CV_LANG_CHECK_HE = "1. LANGUAGE: Is the CV body written in Hebrew? Technical terms (technologies, tools, company names, software names) must stay in English — that is correct. But all prose — profile text, bullet points, job titles, education descriptions — must be in Hebrew. Fix any English prose that slipped through. Do NOT translate the candidate's name."

CV_PASS1_PROMPT = """You are a senior CV writer and recruiter expert.
Create a tailored CV in {language} based on the original CV and job requirements.

MANDATORY WORD-BUDGET PLANNING — do this BEFORE writing any content:
1. Target total length: 480-550 words (Hebrew) or 550-620 words (English) across ALL
   sections combined. This is what fits one printed page at standard CV formatting.
2. Allocate roughly: Profile 35-45 words | Per role: 3-5 bullets, ~12-18 words each |
   Education 20-35 words | Skills: compact list | Languages 10-15 words.
3. If original content exceeds budget, tighten WORDING first (remove filler adjectives,
   merge clauses) — NOT by deleting whole bullets, unless a role has 5+ bullets and you
   cut only the least job-relevant one.
4. NEVER shorten or remove a bullet containing a quantified achievement (%, $, users,
   scale, team size, time saved).
5. If still over budget after tightening wording, prefer trimming older/less relevant
   roles over recent/more relevant ones — never drop a role entirely unless it's clearly
   unrelated filler.

ABSOLUTE RULES — never break these:
1. ONE PAGE MAXIMUM — the final CV must fit on a single printed page, no exceptions. Cut the least relevant bullet points or shorten descriptions to stay within one page. Never sacrifice core tech, key metrics, or the profile to save space — cut filler first.
2. NEVER invent experience, skills, dates, or company names
3. Do not reorder major sections — always: Profile → Experience → Education → Skills → Languages
4. Languages section is ALWAYS last
5. Do NOT present freelance or independent projects as full-time employment positions
6. Do NOT change the chronological order of work experience entries
7. PRESERVE ALL URLs — every URL in the original CV (GitHub, LinkedIn, portfolio, personal site) MUST appear verbatim in the [CONTACT] section. Do NOT drop, shorten, or paraphrase any URL. If the original CV contains tokens of the form [LINK:display|url], copy them exactly as-is into [CONTACT] — do not alter them.

{language_rule}

COMPANY STRUCTURE IS SACRED — never break this:
- Keep every employer as its own separate block with its own company name, job title, and dates
- NEVER merge bullet points from different companies or roles into a single thematic list
- You may reorder bullet points WITHIN a single role to highlight the most relevant ones first
- You may NOT move bullets across roles or companies under any circumstances
- Use `$ ` (dollar-sign + space) as the bullet marker for ALL bullet points — e.g. `$ Built a REST API using FastAPI`
- Do NOT use •, -, *, –, or any other bullet character — only `$ ` at the start of bullet lines

HEADLINE / SENIORITY GUARDRAILS:
- Use the candidate's real core title as the base (e.g. Backend Developer, Software Engineer)
- You may append a focused orientation if supported by real coursework or projects (e.g. Backend Developer | AI & Data Foundations) — keep it subtle and credible
- Do NOT add seniority levels (Senior, Lead, Staff, Principal, etc.) unless the candidate's CV explicitly shows they held such a title
- Do NOT introduce a completely new domain title based on courses alone

PROFILE WRITING RULES:
- Write what the candidate genuinely brings from their real experience
- Do NOT copy or paraphrase sentences from the job description
- The profile must sound like the candidate speaking about themselves
- Make it personal, specific, and grounded in what is actually in the CV

CORE TECH & HIGH-VALUE METRICS — never omit these:
- NEVER remove core programming languages or technologies from the original CV
- SCAN the original CV for outstanding high-value data points: perfect/near-perfect grades, academic honors, scholarships, significant quantitative achievements
- If the target role requires algorithmic, mathematical, or research-heavy skills: surface any strong math background prominently in the PROFILE
- For all other roles: bring relevant quantitative achievements into the PROFILE or top bullet points
- These high-value signals must never be buried or cut

BOLD FORMATTING FOR RECRUITER SCANNING:
- Use **double asterisks** around 3–6 key terms per section — this applies to EVERY section including [PROFILE]
- In [PROFILE]: bold 2–4 role-relevant technologies or measurable skills (e.g. **Python**, **FastAPI**, **95% test coverage**)
- In [EXPERIENCE] bullets: bold the most impactful technology or metric per bullet
- Bold: specific technologies, measurable achievements, and role-critical skills
- Do NOT bold generic words (e.g. "team player", "motivated", "experience")

OUTPUT FORMAT — use these exact section markers. Rules:
- Write each marker on its own line, exactly as shown — no #, no **, no other prefix or suffix
- Immediately follow each marker with the real content on the next line — no blank line between marker and content
- Do NOT write the marker name as a heading or label — it is replaced by the parser

[NAME]
[HEADLINE]
[CONTACT]
[PROFILE]
[EXPERIENCE]
[EDUCATION]
[SKILLS]
[LANGUAGES]
=== ORIGINAL CV ===
{cv_text}
=== CANDIDATE ANSWERS TO QUESTIONS ===
{answers_text}
=== JOB DESCRIPTION ===
{job_text}"""

CV_PASS2_PROMPT = """You are a ruthless senior recruiter reviewing a tailored CV before it goes to a hiring manager.
Review and improve this CV against ALL of these criteria — fix every issue you find:

{language_check}
2. COMPANY STRUCTURE: Is every employer kept as its own separate block? If bullet points from different companies or roles have been merged — restore the original per-company structure immediately.
3. PROFILE AUTHENTICITY: Does the profile sound like the candidate speaking from their real experience? If it echoes the job description's language — rewrite it.
4. SENIORITY CHECK: Remove any Senior/Lead/Staff/Principal or inflated domain title not supported by actual held job titles. A subtle orientation suffix is acceptable if backed by real projects or courses.
5. BOLD FORMATTING: Are 3–6 key terms per section bolded with **double asterisks**? Bold specific technologies, measurable results, and role-critical skills.
6. SPECIFICITY: Remove vague buzzwords. Replace with concrete examples or remove entirely.
7. LENGTH: Must fit ONE page. Cut filler first. Core tech, key metrics, and the profile must be preserved.
8. HUMAN TONE: Should not sound AI-generated. Adjust phrasing if needed.
9. CORE TECH & METRICS PRESERVATION: Are all core programming languages still present? Are high-value data points (grades, honors, achievements) visible? Restore if removed.
10. HONESTY: Do not add anything not in the original CV. Do not present freelance work as full-time employment.
11. URLs: Are all URLs from the ORIGINAL CV present verbatim in [CONTACT]? Restore any missing URLs by copying them exactly from the ORIGINAL CV. If either the original or the draft contains [LINK:display|url] tokens, copy them unchanged — do not unwrap or reformat them.
12. CONTENT COMPLETENESS: Compare the draft's [EXPERIENCE] section against the ORIGINAL CV role by role. If a bullet, technology, or metric from the original was dropped without an equivalent replacement, restore it.

Output ONLY the improved CV using the same section markers.
CRITICAL FORMAT RULES — any violation breaks the Word document:
- Write each marker on its own line exactly: [NAME], [HEADLINE], [CONTACT], [PROFILE], [EXPERIENCE], [EDUCATION], [SKILLS], [LANGUAGES]
- Do NOT prefix markers with #, ##, **, or any other character
- Do NOT add explanations, comments, or labels outside the CV content

=== ORIGINAL CV (ground truth — use this to verify nothing was lost or invented) ===
{original_cv}
=== DRAFT CV TO REVIEW ===
{cv_draft}
=== JOB DESCRIPTION ===
{job_text}"""


# ── Diff-screen constants & prompt ────────────────────────────────────────────

DIFF_SECTIONS = ['[PROFILE]', '[EXPERIENCE]', '[EDUCATION]', '[SKILLS]', '[LANGUAGES]']
SECTION_LABELS = {
    '[PROFILE]':    'פרופיל',
    '[EXPERIENCE]': 'ניסיון תעסוקתי',
    '[EDUCATION]':  'השכלה',
    '[SKILLS]':     'כישורים',
    '[LANGUAGES]':  'שפות',
}

COVER_LETTER_PROMPT = """You are an expert career coach writing a professional cover letter.

Write a concise, targeted cover letter based on the candidate's CV and the job description below.

Requirements:
- Exactly 3 paragraphs (no more, no less)
- Paragraph 1: Open with a confident, specific statement about why this candidate is a strong fit for THIS role — reference 1-2 concrete achievements from the CV
- Paragraph 2: Connect the candidate's most relevant skills/experiences directly to the job's key requirements
- Paragraph 3: Close with enthusiasm and a clear call to action
- Tone: professional, warm, confident — never generic or formulaic
- Length: 180-260 words total
- Language: {language} — write the ENTIRE letter in {language} (including greeting and closing)
- Do NOT include placeholders like [Your Name] or [Date] — write a complete, ready-to-send letter
- Address the hiring team as "Hiring Team" or equivalent in the target language
- Sign off with the candidate's name from the CV

Output ONLY the cover letter text — no explanations, no labels, no markdown.

=== CANDIDATE CV ===
{cv_text}

=== JOB DESCRIPTION ===
{job_text}"""


CV_DIFF_PROMPT = """You are a CV comparison assistant.

I provide you with 5 sections from an AI-ADAPTED CV and the full text of the ORIGINAL CV.
For each section: find the matching content in the ORIGINAL CV, then compare it to the adapted version.

Return ONLY a valid JSON array — no markdown fences, no extra text, just the raw JSON array.

Each element must have exactly these fields:
{{"id":<integer>,"section_name":"<marker>","label":"<Hebrew label>","original_text":"<matched text from ORIGINAL CV — copy verbatim>","updated_text":"<the adapted text provided below>","changed":<true|false>,"explanation_hebrew":"<≤12-word Hebrew sentence — why this helps for the job; empty string if not changed>"}}

Sections from the ADAPTED CV:
{sections_block}

Rules:
- original_text: find the relevant paragraph(s) in the ORIGINAL CV by meaning — copy verbatim from it
- changed: true only when there is a meaningful semantic difference (ignore punctuation/whitespace)
- explanation_hebrew: mention specific keywords or technologies added; empty string if changed is false
- If a section exists in ADAPTED but has no equivalent in ORIGINAL, set original_text to ""

=== ORIGINAL CV ===
{original_cv}

=== JOB CONTEXT (first 200 chars) ===
{job_context}"""


def _resolve_model(model_alias: str) -> str:
    """Map user-facing alias to actual Anthropic model ID."""
    if model_alias == "fable":
        return "claude-fable-5"
    if model_alias == "haiku":
        return "claude-haiku-4-5-20251001"
    return "claude-sonnet-4-6"  # default

# ── Request / Response models ─────────────────────────────────────────────────

class VerifyLicenseRequest(BaseModel):
    licenseKey: str

class AnalyzeRequest(BaseModel):
    licenseKey: Optional[str] = None
    cvText: str
    jobText: str
    answers: list = []
    preflight: bool = False
    model: str = "claude-sonnet-4-6"

class GenerateCVRequest(BaseModel):
    licenseKey: Optional[str] = None
    cvText: str
    jobText: str
    jobLanguage: str = "english"
    answers: list[dict] = []
    cvUrls: list[str] = []
    userConstraints: str = ""
    generateCoverLetter: bool = False
    enableTracking: bool = True
    jobTitle: str = ""
    company: str = ""
    model: str = "sonnet"

class RankJobsRequest(BaseModel):
    licenseKey: Optional[str] = None
    cvText: str
    jobs: list[dict]

async def _rank_single_job(cv_summary: str, job: dict) -> Optional[dict]:
    """Score one job against the CV with a focused single-job prompt. Returns None on failure."""
    title = job.get("title") or "Unknown"
    url = job.get("url") or ""
    text = (job.get("text") or "")[:2000]
    job_block = f"{title}\n{url}\n\n{text}"
    prompt = RANK_SINGLE_JOB_PROMPT.format(cv_summary=cv_summary, job_text=job_block)
    try:
        raw = await call_claude(prompt, max_tokens=120)
        item = parse_json_response(raw)
        if isinstance(item, dict) and "score" in item:
            return {
                "job": job,
                "score": max(0, min(100, int(item["score"]))),
                "pro": str(item.get("pro", "")),
                "con": str(item.get("con", "")),
            }
    except Exception:
        pass
    return None


class ScrapeJobRequest(BaseModel):
    url: str
    text: str
    title: str = ""

class ImportJobsRequest(BaseModel):
    cvText: str
    minScore: int = 70
    timeRange: str = "3days"  # "3days" | "since_last"

RANK_SINGLE_JOB_PROMPT = """You are a strict but fair senior hiring manager. Score this single job posting against the candidate's CV.

STEP 1 — classify every requirement:
  CRITICAL = "must", "required", "mandatory", or listed in the main requirements section
  SECONDARY = "nice to have", "preferred", "advantage", "bonus", or in a lower-priority section

SCORING — start from 100 and deduct:
  Missing CRITICAL requirement: -15 to -25 per item (proportional to centrality)
  Partial/theoretical on a CRITICAL item: -5 to -12
  Years shortfall: (required - actual) / required × 30 pts
  Seniority mismatch (Senior/Lead in title but not in CV): cap at 65
  Missing SECONDARY / nice-to-have: -2 to -5 per item (never more than -5 each)
Positive offsets (secondary gaps only): strong academics +5, adjacent skills +5, relevant projects +5

DOMAIN MISMATCH — cap at 55 ONLY when the candidate genuinely lacks the critical skills. If they have the required skills but come from a different job-title background, do NOT cap — score on skill fit alone.

CALIBRATION: 85+ shortlist | 70-84 interview | 55-69 real gaps | <40 wrong fit
A candidate meeting all CRITICAL requirements but lacking all SECONDARY ones → 65-75, not below 55.
A score below 50 means most CRITICAL requirements are missing — not just a different job background.

Return ONLY valid JSON — no markdown, no explanation:
{{"score": <0-100>, "pro": "<max 10 words — strongest match signal>", "con": "<max 10 words — biggest gap>"}}

=== CANDIDATE CV ===
{cv_summary}

=== JOB POSTING ===
{job_text}"""

RANK_JOBS_PROMPT = """You are a strict but fair senior engineering hiring manager.
Score how well THIS candidate's CV matches each job listing. Be honest and realistic.

STEP 1 — READ THE JOB TEXT AND CLASSIFY EACH REQUIREMENT:
- "required" / "must" / "mandatory" / listed first → CRITICAL
- "preferred" / "advantage" / "nice to have" / listed later → SECONDARY
- When the job is vague, infer criticality from position and emphasis in the text

STEP 2 — SCORE BASED ON ACTUAL FIT:
Start from 100 and deduct based on what is missing:

CRITICAL gaps (each deduction is proportional to how critical the requirement appears):
- Years of experience shortfall: required N yrs, CV shows M yrs → deduct ~(N-M)/N × 30 pts
  (e.g. 4yr req, 2yr CV → -15; 5yr req, 0yr CV → -30)
- Missing core tech listed as "required": -15 to -25 per item depending on how central it is
- Partial/theoretical match on a required skill: -5 to -12
- Seniority mismatch: Senior/Lead/Staff title with no evidence of that level in CV → cap score at 65
- Domain mismatch: cap score at 55 ONLY when the candidate genuinely lacks the critical skills. If they possess the required skills but come from a different job-title background, do NOT apply this cap — score purely on skill fit. A backend developer who has SQL + BI tools + ETL knowledge is NOT a domain mismatch for a BI role.

SECONDARY gaps: -2 to -5 per missing item (NEVER more than -5 per item)
IMPORTANT: A candidate who meets all CRITICAL requirements but lacks all SECONDARY ones should score 65-75, not below 55.
A score below 50 means most CRITICAL requirements are absent — not that the job title on the CV sounds different.

POSITIVE signals that can partially offset secondary gaps (not critical ones):
- Strong relevant academic background: +5
- Closely adjacent skills that transfer: +5
- Relevant projects or open source: +5

CALIBRATION:
- 85+: Strong match — would shortlist immediately
- 70-84: Good match — worth an interview despite minor gaps
- 55-69: Partial match — real gaps exist
- 40-54: Significant blocker — not ready for this role level
- <40: Wrong domain or severe experience shortfall

OUTPUT — return ONLY a valid JSON array, no markdown, no explanation:
[{{"index":0,"score":72,"pro":"...","con":"..."}}]
- pro: one sentence max 10 words — strongest signal in this candidate's favor
- con: one sentence max 10 words — single most serious gap for THIS specific job

=== CANDIDATE CV ===
{cv_summary}

=== JOBS TO RANK ===
{jobs_list}"""


# ── App ───────────────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    if not ANTHROPIC_API_KEY:
        print("WARNING: ANTHROPIC_API_KEY is not set")
    if not GUMROAD_PRODUCT_PERMALINK:
        print("WARNING: GUMROAD_PRODUCT_PERMALINK is not set")
    if not GUMROAD_ACCESS_TOKEN:
        print("INFO: GUMROAD_ACCESS_TOKEN not set — seller-side API features disabled")
    print(f"[JMA:startup] product={GUMROAD_PRODUCT_PERMALINK} upgrade_url={UPGRADE_URL}")
    yield

app = FastAPI(title="Job Match AI Server", lifespan=lifespan)

# ── CORS — raw ASGI middleware ────────────────────────────────────────────────
# We use a raw ASGI middleware class instead of Starlette's CORSMiddleware or
# @app.middleware("http").  Both of those go through BaseHTTPMiddleware which,
# in some Starlette versions, intercepts WebSocket upgrade requests and can send
# a rejection before websocket.accept() is ever called.  A raw ASGI class that
# explicitly passes scope["type"] == "websocket" straight through is the only
# reliable way to keep WebSocket connections unaffected by CORS logic.
_CORS_HEADERS: list[tuple[bytes, bytes]] = [
    (b"access-control-allow-origin",  b"*"),
    (b"access-control-allow-methods", b"GET, POST, PUT, DELETE, OPTIONS"),
    (b"access-control-allow-headers", b"*"),
]


class _CORSMiddleware:
    """Adds CORS headers to all HTTP responses; WebSocket scopes pass through untouched."""

    def __init__(self, app: ASGIApp) -> None:
        self._app = app

    async def __call__(self, scope: Scope, receive: Receive, send: Send) -> None:
        # WebSocket or lifespan — pass directly to the app, zero interference.
        if scope["type"] != "http":
            await self._app(scope, receive, send)
            return

        # Preflight — short-circuit with 200 + CORS headers.
        if scope.get("method") == "OPTIONS":
            await send({
                "type": "http.response.start",
                "status": 200,
                "headers": _CORS_HEADERS + [(b"access-control-max-age", b"86400")],
            })
            await send({"type": "http.response.body", "body": b""})
            return

        # All other HTTP — inject CORS headers into whatever the app returns.
        async def send_with_cors(message: dict) -> None:
            if message["type"] == "http.response.start":
                existing = list(message.get("headers", []))
                message = {**message, "headers": existing + _CORS_HEADERS}
            await send(message)

        await self._app(scope, receive, send_with_cors)


app.add_middleware(_CORSMiddleware)

from app.routes.points import router as points_router  # noqa: E402  (after app/_ws_user_id defined — deps.py imports main lazily)
app.include_router(points_router)

from app.routes.recruiters import router as recruiters_router  # noqa: E402
app.include_router(recruiters_router)


@app.get("/health")
async def health():
    return {"status": "ok"}


# ── CV Profile Extraction ─────────────────────────────────────────────────────
# Architecture: static system prompt is cached (cache_control: ephemeral) so the
# ~1 200-token instruction block is billed only on the first call per cache window.
# The dynamic CV text travels in the user message, which is never cached.



EXTRACT_PROFILE_SYSTEM = """\
You are an expert tech-recruiter talent analyst. Extract a structured JSON profile \
from the candidate CV the user will provide. Follow every instruction exactly.

OUTPUT: Return ONLY valid JSON — no markdown, no explanation, no extra keys.

JSON SCHEMA (fill all fields; use 0 / false / [] / {} when data is absent):
{
  "industry_summary": {
    "total_years_industry": <float>,
    "domain_years": {
      "<domain>": <float>
      // only domains actually found: backend, frontend, fullstack, mobile, devops_cloud,
      // ai_ml_llm, data_bi, cyber, qa, embedded, other
    }
  },
  "traits": [<string>],
  "experience": {
    "backend":      { "<tech>": {"industry_years": <float>, "personal_years": <float>, "personal_weight": <int 0-100>, "search_tags": [<string>]} },
    "frontend":     { ... same shape ... },
    "ai_ml_llm":    { ... },
    "data_bi":      { ... },
    "devops_cloud": { ... },
    "mobile":       { ... },
    "other_domains":{ ... }
  },
  "tools_and_methods": { "<tool>": {"found": <bool>, "search_tags": [<string>]} },
  "languages": { "<lang>": "<level>" }
}

CRITICAL RULES:
1. SEPARATION: `industry_years` = verified paid employment only. \
`personal_years` = side projects, academia, bootcamps, online courses, open source.
2. YEAR-TO-TECH ATTACHMENT: Attach years to a tech ONLY from roles/projects whose \
description actually mentions that tech (or an unambiguous alias). NEVER spread total \
career years across all listed skills. A skill that appears only in a "Skills" list \
with no role/project context gets: industry_years 0, personal_years 0.5, personal_weight 25.
3. PERSONAL EXPERIENCE WEIGHTING: For every tech, set `personal_weight` (0-100) = how \
substantial the NON-industry experience is, judged from CV evidence:
   - 90-100: production system with real users; launched startup MVP; paid freelance; \
     open-source with real adoption
   - 60-85:  end-to-end project that was actually deployed/live; working product built \
     independently; meaningful contribution to a real codebase
   - 35-55:  substantial portfolio project; final degree project; hackathon finalist work
   - 10-30:  coursework, bootcamp exercises, tutorials, "familiar with X" claims
   - personal_years == 0 → personal_weight = 0. No evidence either way → 40.
   Signals of substance: "deployed", "production", "live", "users", scale numbers, \
   links to a working product, App Store / cloud hosting mentions, depth of description. \
   One-line mention = low. Detailed project with outcomes = high.
4. DURATION INFERENCE: If a tech appears inside a role with explicit dates, compute its \
duration from those dates. If dates are missing, default to 1.0 year.
5. TRAITS: English adjective/noun phrases only, e.g. "Analytically-Minded", "Team-Player".
6. DOMAIN YEARS in `industry_summary.domain_years` = sum of `industry_years` across all \
techs in that domain, capped at total career years.
7. FULLSTACK INFERENCE: If ≥1 industry year in both backend AND frontend, add \
"fullstack" with min(backend_years, frontend_years).
8. DOMAIN PLACEMENT: Place each tech in the domain matching HOW THE CANDIDATE USED IT \
(Python for ML pipelines → ai_ml_llm). Never duplicate a tech across domains.
9. `tools_and_methods`: include git, cicd, agile, scrum, docker, kubernetes, \
microservices, rest, graphql, tdd, and any others found.

SEARCH TAGS RULES (apply to every tech/tool entry):
Populate `search_tags` with synonyms, aliases, and equivalents for raw text matching. \
LIMIT: max 6 per entry. Every Latin-script tag MUST be at least 3 characters long \
(never emit tags like "go", "ai", "bi", "ml" — use "golang", "genai", "power bi", \
"machine learning" instead). Include the best from:
a. Industry-standard English synonyms/acronyms ("PostgreSQL" → ["SQL", "Postgres", "RDBMS"]).
b. The most common Hebrew transliteration used in Israeli job postings, including at \
least one prefix-attached form ("Python" → ["פייתון", "בפייתון"]).
c. One related higher-level domain if highly relevant ("React" → ["Frontend"]).
All Hebrew must be clean UTF-8 (no \\uXXXX escapes)."""
# You are an expert tech-recruiter talent analyst. Extract a structured JSON profile \
# from the candidate CV the user will provide. Follow every instruction exactly.

# OUTPUT: Return ONLY valid JSON — no markdown, no explanation, no extra keys.

# JSON SCHEMA (fill all fields; use 0 / false / [] / {} when data is absent):
# {
#   "industry_summary": {
#     "total_years_industry": <float>,
#     "domain_years": {
#       "<domain>": <float>
#       // only domains actually found: backend, frontend, fullstack, mobile, devops_cloud,
#       // ai_ml_llm, data_bi, cyber, qa, embedded, other
#     }
#   },
#   "traits": [<string>],
#   "experience": {
#     "backend":      { "<tech>": {"industry_years": <float>, "personal_years": <float>, "search_tags": [<string>]} },
#     "frontend":     { "<tech>": {"industry_years": <float>, "personal_years": <float>, "search_tags": [<string>]} },
#     "ai_ml_llm":    { "<tech>": {"industry_years": <float>, "personal_years": <float>, "search_tags": [<string>]} },
#     "data_bi":      { "<tech>": {"industry_years": <float>, "personal_years": <float>, "search_tags": [<string>]} },
#     "devops_cloud": { "<tech>": {"industry_years": <float>, "personal_years": <float>, "search_tags": [<string>]} },
#     "mobile":       { "<tech>": {"industry_years": <float>, "personal_years": <float>, "search_tags": [<string>]} },
#     "other_domains":{ "<tech>": {"industry_years": <float>, "personal_years": <float>, "search_tags": [<string>]} }
#   },
#   "tools_and_methods": { "<tool>": {"found": <bool>, "search_tags": [<string>]} },
#   "languages": { "<lang>": "<level>" }
# }

# CRITICAL RULES:
# 1. SEPARATION: `industry_years` = verified paid employment only. \
# `personal_years` = side projects, academia, bootcamps, online courses.
# 2. DURATION INFERENCE: If a tech is listed inside a job role that has explicit dates, \
# compute its duration from those dates. If dates are missing, default to 1.0 year.
# 3. TRAITS: English adjective/noun phrases only. Examples: "Analytically-Minded", \
# "Self-Directed", "Team-Player".
# 4. DOMAIN YEARS in `industry_summary.domain_years` = sum of `industry_years` across all \
# techs in that domain, capped at total career years.
# 5. FULLSTACK INFERENCE: If the person has ≥1 year backend AND ≥1 year frontend \
# (industry), add "fullstack" to domain_years with min(backend_years, frontend_years).
# 6. `tools_and_methods`: include git, cicd, agile, scrum, docker, kubernetes, \
# microservices, rest, graphql, tdd, and any others found.

# SEARCH TAGS RULES (apply to every tech/tool entry):
# For each extracted skill, technology, framework, or qualification populate `search_tags` \
# with synonyms, aliases, and equivalents for raw text pattern-matching. \
# LIMIT: maximum 6 tags per entry — choose only the most useful ones. Include the best from:
# a. Industry-standard English synonyms and acronyms \
#    (e.g. "PostgreSQL" → ["SQL", "Postgres", "DB", "RDBMS"]).
# b. The most common Hebrew transliteration/translation used in Israeli job postings. \
#    CRITICAL: also include the most frequent prefix-attached forms because Hebrew prepends \
#    prepositions directly to words — add at least one prefixed variant for each Hebrew tag \
#    (e.g. "Python" → ["פייתון", "בפייתון"]; \
#     "Machine Learning" → ["למידת מכונה", "בלמידת מכונה"]; \
#     "Database" → ["בסיסי נתונים", "דאטהבייס"]).
# c. One related higher-level domain if highly relevant \
#    (e.g. "React" → ["Frontend"]).
# All Hebrew characters must be output as valid, clean UTF-8 strings inside the JSON \
# (do NOT escape them as \\uXXXX sequences)."""


class ExtractProfileRequest(BaseModel):
    cvText: str

@app.post("/api/extract-profile")
async def extract_profile(body: ExtractProfileRequest, x_license_key: Optional[str] = Header(None)):
    """Extract a structured skills/experience profile from raw CV text using Claude."""
    license_key = x_license_key or ""
    try:
        await verify_gumroad_license(license_key)
    except Exception:
        raise HTTPException(status_code=403, detail="Invalid license")

    if not body.cvText or len(body.cvText.strip()) < 50:
        raise HTTPException(status_code=422, detail="CV text too short")

    system_blocks = [
        {
            "type": "text",
            "text": EXTRACT_PROFILE_SYSTEM + "\n\nIMPORTANT: Return ONLY raw JSON. No conversational text, no markdown wrappers like ```json.",
            "cache_control": {"type": "ephemeral"},
        }
    ]
    user_content = f"=== CANDIDATE CV ===\n{body.cvText[:6000]}\n=== END CV ==="

    try:
        # שימוש במזהה הרשמי המדויק מתוך טבלת הדוקומנטציה שלך (image_c410de.png)
        raw, stop = await call_claude_cached(
            system_blocks=system_blocks,
            user_content=user_content,
            max_tokens=8192,
            model="claude-sonnet-4-6",  # מבוסס על ה-Claude API ID מהטבלה שלך
        )
        if stop == "max_tokens":
            print("[JMA:extract_profile] truncated (max_tokens) — attempting partial parse")
            
        # רשת ביטחון: אם המודל החדש החזיר טיפוס מורכב (Union) ולא מחרוזת נקייה, נחלץ את הטקסט בבטחה
        if hasattr(raw, "text"):
            raw_text = raw.text
        elif isinstance(raw, list) and len(raw) > 0 and hasattr(raw[0], "text"):
            raw_text = raw[0].text
        else:
            raw_text = str(raw)

        profile = parse_json_response(raw_text)
        
        if not isinstance(profile, dict) or "experience" not in profile:
            raise ValueError("Invalid profile shape")
            
        from fastapi.responses import JSONResponse
        return JSONResponse(content={"profile": profile}, media_type="application/json; charset=utf-8")
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[JMA:extract_profile] {type(e).__name__}: {e}")
        raise HTTPException(status_code=500, detail="Profile extraction failed")

# @app.post("/api/extract-profile")
# async def extract_profile(body: ExtractProfileRequest, x_license_key: Optional[str] = Header(None)):
#     """Extract a structured skills/experience profile from raw CV text using Claude."""
#     license_key = x_license_key or ""
#     try:
#         await verify_gumroad_license(license_key)
#     except Exception:
#         raise HTTPException(status_code=403, detail="Invalid license")

#     if not body.cvText or len(body.cvText.strip()) < 50:
#         raise HTTPException(status_code=422, detail="CV text too short")

#     system_blocks = [
#         {
#             "type": "text",
#             "text": EXTRACT_PROFILE_SYSTEM,
#             "cache_control": {"type": "ephemeral"},
#         }
#     ]
#     user_content = f"=== CANDIDATE CV ===\n{body.cvText[:6000]}\n=== END CV ==="

#     try:
#         raw, stop = await call_claude_cached(
#             system_blocks=system_blocks,
#             user_content=user_content,
#             max_tokens=8192,
#             model="claude-sonnet-4-6",
#         )
#         if stop == "max_tokens":
#             print("[JMA:extract_profile] truncated (max_tokens) — attempting partial parse")
#         profile = parse_json_response(raw)
#         if not isinstance(profile, dict) or "experience" not in profile:
#             raise ValueError("Invalid profile shape")
#         from fastapi.responses import JSONResponse
#         return JSONResponse(content={"profile": profile}, media_type="application/json; charset=utf-8")
#     except HTTPException:
#         raise
#     except Exception as e:
#         print(f"[JMA:extract_profile] {type(e).__name__}: {e}")
#         raise HTTPException(status_code=500, detail="Profile extraction failed")


# class QuickScoreRequest(BaseModel):
#     cvText: str = ""
#     jobText: str = ""


# @app.post("/api/quick-score")
# async def quick_score(body: QuickScoreRequest):
#     """Ultra-fast match percentage + short reason bullets. Uses cheapest model, no auth."""
#     if not body.cvText or not body.jobText:
#         return {"score": 0, "bullets": []}
#     he_chars = sum(1 for c in body.jobText[:300] if "א" <= c <= "ת")
#     lang = "Hebrew" if he_chars > 20 else "English"
#     prompt = (
#         f"Score how well this CV matches this job (0–100). Reply JSON only.\n\n"
#         f"CV:\n{body.cvText[:900]}\n\nJob:\n{body.jobText[:900]}\n\n"
#         f'Reply: {{"score": <0-100>, "bullets": ["✅ short reason", "❌ short reason", "✅ short reason"]}}\n'
#         f"3–4 bullets, each under 35 chars, in {lang}. No extra text."
#     )
#     try:
#         resp = await anthropic_client.messages.create(
#             model="claude-haiku-4-5-20251001",
#             max_tokens=110,
#             messages=[{"role": "user", "content": prompt}],
#         )
#         data = _safe_json_loads(resp.content[0].text.strip())
#         return {
#             "score": max(0, min(100, int(data.get("score", 0)))),
#             "bullets": (data.get("bullets") or [])[:4],
#         }
#     except Exception as e:
#         print(f"[JMA:quick_score] {e}")
#         return {"score": 0, "bullets": []}


class AnalyzeStreamRequest(BaseModel):
    cvText: str = ""
    jobText: str = ""
    answers: list[dict] = []
    userConstraints: str = ""
    model: str = "claude-sonnet-4-6"


class StreamQuestionsRequest(BaseModel):
    cvText:     str = ""
    jobText:    str = ""
    model:      str = "claude-sonnet-4-6"
    baseScore:  int = -1
    gapPct:     int = -1


STREAM_ANALYSIS_PROMPT = """\
You are a senior career consultant. The candidate CV is in your system prompt.

Job Description:
{job_text}
{answers_block}{constraints_block}

Write a concise, direct, actionable match analysis in {lang}. Cover:
(1) overall match verdict, (2) strongest alignment points, (3) key gaps to address,
(4) specific recommendations.
Use short paragraphs. Max 350 words.

IMPORTANT — emit output iteratively: write and flush each paragraph as soon as it is complete. \
Do NOT buffer the entire response before outputting."""

STREAM_QUESTIONS_PROMPT = """\
You are a senior career consultant. The candidate CV is in your system prompt.

The initial match score is {base_score}% with {gap_pct} improvement points available.
{answers_block}
Identify the {n_questions} most important skill gaps that are unclear or missing in the CV \
and clearly required by the job below.

CRITICAL STREAMING RULE — output each question the MOMENT you identify it, one at a time:
1. Identify gap #1 → immediately output its JSON object on its own line.
2. Identify gap #2 → immediately output its JSON object on its own line.
Continue until all {n_questions} gaps are output.

For EACH gap output exactly one line of valid JSON (no markdown, no wrapper array):
{{"id":"q1","skill":"<English name>","question":"<מקצועית קצרה>","explanation":"<עד 15 מילים בעברית>","weight":<integer>}}

The integer weights MUST sum to exactly {gap_pct}.
ALL Hebrew text fields must be in Hebrew. Skill names stay in English.

After the last question output exactly:
[QUESTIONS_DONE]

=== JOB DESCRIPTION ===
{job_text}"""


# ── Live-streaming questions prompt ───────────────────────────────────────────
# Same question logic / content as STREAM_QUESTIONS_PROMPT above.
# Output format changed so the question text can be streamed token-by-token:
#   [META]{"id":"q1","skill":"Python","weight":8}   ← emitted immediately
#   <full Hebrew question text on one line>          ← streamed word by word
#   [EXPL]<explanation up to 15 Hebrew words>[/EXPL] ← closes the question
#
# The parser (_stream_q_parse) detects these markers and emits q_open / q_token
# / q_close SSE events so the client can open a textarea before the sentence ends.

# STREAM_QUESTIONS_LIVE_PROMPT = """\
# You are a senior career consultant. The candidate CV is in your system prompt.

# The initial match score is {base_score}% with {gap_pct} improvement points available.

# Identify the {n_questions} most important skill gaps that are unclear or missing in \
# the CV and are clearly required by the job below.

# CRITICAL OUTPUT FORMAT — for EACH gap emit EXACTLY these 3 lines the MOMENT you identify it:
# Line 1 (metadata — emit immediately): [META]{{"id":"qN","skill":"<English name>","weight":<int>}}
# Line 2 (question   — stream freely):  <full Hebrew question text, one line>
# Line 3 (close):                        [EXPL]<Hebrew explanation, max 15 words>[/EXPL]

# Rules:
# • Your FIRST token must be `[` — no preamble, no thinking text, no introduction.
# • Emit gap #1 the instant you find it — do NOT read the whole job first.
# • weights must sum to exactly {gap_pct}. Skill names in English; all text in Hebrew.
# • Do NOT add any other lines, commentary, or JSON wrappers.
# • After the last question output: [QUESTIONS_DONE]

# === JOB DESCRIPTION ===
# {job_text}"""

STREAM_QUESTIONS_LIVE_PROMPT = """\
You are a senior career consultant. The candidate CV is in your system prompt.

The initial match score is {base_score}% with {gap_pct} improvement points available.

Identify UP TO {n_questions} of the most critical and substantial skill gaps that are unclear or missing in \
the CV and are clearly required by the job below. Focus only on high-impact gaps; do not force minor questions just to meet a count.

CRITICAL OUTPUT FORMAT — for EACH gap emit EXACTLY these 3 lines the MOMENT you identify it:
Line 1 (metadata — emit immediately): [META]{{"id":"qN","skill":"<English name>","weight":<int>}}
Line 2 (question   — stream freely):  <full Hebrew question text, one line>
Line 3 (close):                        [EXPL]<Hebrew explanation, max 15 words>[/EXPL]

Rules:
• Your FIRST token must be `[` — no preamble, no thinking text, no introduction.
• Emit gap #1 the instant you find it — do NOT read the whole job first.
• weights must sum to exactly {gap_pct}. Skill names in English; all text in Hebrew.
• Do NOT add any other lines, commentary, or JSON wrappers.
• After the last question output: [QUESTIONS_DONE]

--- ADVANCED GAP EVALUATION RULES ---
1. DO NOT ask novice-level questions if the CV demonstrates advanced experience in a related domain. 
2. If the job requires AI tooling (like Copilot/Cursor) and the CV already shows hands-on AI Development (LLMs, Prompts, Agents), acknowledge their AI background in the question rather than assuming they lack AI literacy.
3. Tailor the phrasing to match the candidate's existing professional depth (e.g., blend their Java/Python/AI background into the context of the question).

=== JOB DESCRIPTION ===
{job_text}"""

async def _stream_q_parse(token_iter):
    """
    3-state parser for STREAM_QUESTIONS_LIVE_PROMPT output.
    Yields SSE-formatted data strings (not full 'data: ...' lines yet — caller wraps).

    States:
      WAIT_META  → accumulate until [META]...\\n is complete → emit q_open
      STREAM_TEXT → flush tokens immediately (keeping last GUARD chars) → on [EXPL] emit final text
      COLLECT_EXPL → accumulate until [/EXPL] → emit q_close → back to WAIT_META
    """
    state  = "WAIT_META"
    buf    = ""
    q_id   = ""
    GUARD  = 14          # bytes kept back to detect split markers

    async for chunk in token_iter:
        buf += chunk

        # Inner loop: drain buf through state transitions without waiting for next chunk
        changed = True
        while changed:
            changed = False

            if state == "WAIT_META":
                if "[META]" in buf:
                    start  = buf.index("[META]")
                    after  = buf[start + 6:]
                    if "\n" in after:
                        meta_raw, buf = after.split("\n", 1)
                        try:
                            meta  = json.loads(repair_json(meta_raw.strip()))
                            q_id  = meta.get("id", "")
                            yield f"data: {json.dumps({'q_open': meta})}\n\n"
                            state   = "STREAM_TEXT"
                            changed = True
                        except Exception:
                            buf = after  # skip bad meta, keep trying
                    else:
                        buf = buf[start:]   # keep [META]... for next chunk
                elif "[QUESTIONS_DONE]" in buf:
                    return
                else:
                    buf = buf[-GUARD:]  # keep tail in case [META] spans chunks

            elif state == "STREAM_TEXT":
                if "[EXPL]" in buf:
                    idx      = buf.index("[EXPL]")
                    tail_txt = buf[:idx]
                    buf      = buf[idx + 6:]
                    if tail_txt:
                        yield f"data: {json.dumps({'q_token': {'id': q_id, 'text': tail_txt}})}\n\n"
                    state   = "COLLECT_EXPL"
                    changed = True
                elif "[QUESTIONS_DONE]" in buf:
                    pre = buf[:buf.index("[QUESTIONS_DONE]")].strip()
                    if pre:
                        yield f"data: {json.dumps({'q_token': {'id': q_id, 'text': pre}})}\n\n"
                    return
                else:
                    # Flush safe portion; retain GUARD chars for split-marker detection
                    if len(buf) > GUARD:
                        safe = buf[:-GUARD]
                        yield f"data: {json.dumps({'q_token': {'id': q_id, 'text': safe}})}\n\n"
                        buf = buf[-GUARD:]

            elif state == "COLLECT_EXPL":
                if "[/EXPL]" in buf:
                    expl, buf = buf.split("[/EXPL]", 1)
                    yield f"data: {json.dumps({'q_close': {'id': q_id, 'explanation': expl.strip()}})}\n\n"
                    state   = "WAIT_META"
                    changed = True
                elif "[QUESTIONS_DONE]" in buf:
                    return


@app.post("/api/stream-questions")
async def stream_questions_endpoint(
    body: StreamQuestionsRequest,
    x_license_key: Optional[str] = Header(None),
):
    """
    Single streaming call that:
      1. Runs Pass 1 (base score + analysis) — fast non-streaming call.
      2. Emits {meta: {base_score, gap_pct, jobTitle, company, summary}} as first SSE event.
      3. Streams questions token-by-token via q_open / q_token / q_close events.
      4. Emits [DONE].
    The client can open a textarea the moment the first q_open arrives,
    letting the user type while subsequent tokens / questions are still streaming.
    """
    def _sse_err(msg: str):
        async def _gen():
            yield f'data: {json.dumps({"error": msg})}\n\n'
            yield "data: [DONE]\n\n"
        return StreamingResponse(_gen(), media_type="text/event-stream",
                                 headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    if not body.cvText or not body.jobText:
        return _sse_err("חסרים נתוני CV או משרה.")

    license_key = x_license_key or ""
    try:
        await verify_gumroad_license(license_key)
    except Exception:
        return _sse_err("רישיון לא תקף.")

    cv_text  = body.cvText[:3000]
    job_text = body.jobText[:2500]
    sys_blocks     = _cv_system_blocks(cv_text)
    resolved_model = _resolve_model(body.model)

    async def generate():
        # ── Pass 1: base score (skip if client already has a local score) ─────
        base_score = 50
        gap_pct    = 20
        meta_out   = {}
        if body.baseScore >= 0:
            # Client provided local matcher score — skip AI Pass 1 entirely
            base_score = max(0, min(100, body.baseScore))
            gap_pct    = max(0, min(40,  body.gapPct if body.gapPct >= 0 else 20))
            meta_out   = {"base_score": base_score, "gap_pct": gap_pct}
            print(f"[JMA:stream_q] pass1 skipped (client score={base_score} gap={gap_pct})")
        else:
            try:
                raw1, stop1 = await call_claude_cached(
                    system_blocks=sys_blocks,
                    user_content=BASE_ANALYSIS_USER.format(
                        scoring_rules=_SCORING_RULES,
                        job_text=job_text,
                    ),
                    max_tokens=1800,
                    model="claude-3-5-haiku",
                )
                if stop1 != "max_tokens":
                    a = parse_json_response(raw1)
                    base_score = max(0, min(100, int(a.get("base_score", 50))))
                    gap_pct    = max(0, min(40,  int(a.get("gap_pct",    20))))
                    reasoning  = a.get("scoring_reasoning", "")
                    if reasoning:
                        print(f"[JMA:stream_q] scoring_reasoning:\n{reasoning}")
                    meta_out = {
                        "base_score":  base_score,
                        "gap_pct":     gap_pct,
                        "jobTitle":    a.get("jobTitle",    ""),
                        "company":     a.get("company",     ""),
                        "jobLanguage": a.get("jobLanguage", "english"),
                        "summary":     a.get("summary",     ""),
                        "strengths":   a.get("strengths",   []),
                        "hard_gaps":   a.get("hard_gaps",   []),
                    }
                    print(f"[JMA:stream_q] pass1 base={base_score} gap={gap_pct}")
            except Exception as e:
                print(f"[JMA:stream_q] pass1 error: {e}")

        yield f"data: {json.dumps({'meta': meta_out})}\n\n"

        # ── Pass 2: stream questions token-by-token ──────────────────────────
        if gap_pct > 0:
            n_q = 4 if gap_pct >= 20 else 3
            prompt = STREAM_QUESTIONS_LIVE_PROMPT.format(
                base_score=base_score,
                gap_pct=gap_pct,
                n_questions=n_q,
                job_text=job_text,
            )
            try:
                async with anthropic_client.messages.stream(
                    model="claude-sonnet-4-6",
                    max_tokens=700,
                    system=sys_blocks,
                    messages=[{"role": "user", "content": prompt}],
                ) as stream:
                    async for sse_chunk in _stream_q_parse(stream.text_stream):
                        yield sse_chunk
            except Exception as e:
                print(f"[JMA:stream_q] pass2 error: {e}")

        yield "data: [DONE]\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


DEEP_ANALYSIS_PROMPT = """\
You are a senior career consultant. The candidate CV is in your system prompt.

You receive: the job description, the candidate's answers to gap questions, and a
bank of verified answers from PREVIOUS applications. Treat bank facts as reliable
CV supplements.

SCORING METHOD - score ONLY by coverage of the job's actual requirements:
1. Identify the job's core requirements. For each: is it covered by the CV, the
   answers, or the answer bank?
2. Judge answer CONTENT only - NEVER style, length, or phrasing. A short answer
   stating a fact counts as that fact. FORBIDDEN: describing any answer as vague,
   unprofessional, or insufficient.
3. A live MVP, deployed product, or production project counts as significant
   experience even if not employment.
4. Do NOT default to a "safe" 60-65. Requirements substantially covered -> 80+.
   Core must-haves truly absent -> below 40. The score must follow the evidence.

STRICT OUTPUT FORMAT - no preamble, no thinking:
Line 1: [SCORE]<integer 0-100>[/SCORE]
Then a flowing Hebrew analysis (~130 words):
- חוזקות: 2-3 נקודות קונקרטיות מול דרישות המשרה
- פערים אמיתיים בלבד: דרישות ליבה שאינן מכוסות (אם אין - כתוב שאין)
- המלצה חד-משמעית: להגיש / להגיש עם חיזוק קו"ח / לא להגיש - ונימוק במשפט

All text in Hebrew. No markdown. No JSON.

=== JOB DESCRIPTION ===
{job_text}

=== CANDIDATE ANSWERS TO GAP QUESTIONS ===
{answers_text}

=== VERIFIED ANSWER BANK FROM PREVIOUS APPLICATIONS ===
{answer_bank}"""


class DeepAnalysisRequest(BaseModel):
    cvText:     str  = ""
    jobText:    str  = ""
    answers:    list = []
    answerBank: list = []


@app.post("/api/stream-deep-analysis")
async def stream_deep_analysis(body: DeepAnalysisRequest, x_license_key: Optional[str] = Header(None)):
    cv_text  = (body.cvText  or "")[:6000]
    job_text = (body.jobText or "")[:3000]

    answers_text = "\n".join(
        f"שאלה {i+1} ({a.get('skill','?')}, משקל {a.get('weight',0)}%): "
        f'תשובה: "{a.get("answer","לא ענה")}" '
        f'(ציון עצמי: {a.get("sliderValue",0)}%)'
        for i, a in enumerate(body.answers or [])
    ) or "לא נענו שאלות"

    answer_bank = "\n".join(
        f"- {b.get('skill','?')}: \"{b.get('answer','')}\" (ממשרה: {b.get('jobTitle','')})"
        for b in (body.answerBank or [])[:40]
    ) or "אין מאגר תשובות קודמות"

    sys_blocks = _cv_system_blocks(cv_text)
    prompt     = DEEP_ANALYSIS_PROMPT.format(job_text=job_text, answers_text=answers_text,
        answer_bank=answer_bank)

    async def generate():
        import re
        buf        = ""
        score_sent = False

        async with anthropic_client.messages.stream(
            model="claude-haiku-4-5-20251001",
            max_tokens=600,
            system=sys_blocks,
            messages=[{"role": "user", "content": prompt}],
        ) as stream:
            async for token in stream.text_stream:
                if not score_sent:
                    buf += token
                    if "[/SCORE]" in buf:
                        m = re.search(r'\[SCORE\](\d+)\[/SCORE\]', buf)
                        score = max(0, min(100, int(m.group(1)))) if m else 50
                        yield f"data: {json.dumps({'score': score})}\n\n"
                        rest = buf.split("[/SCORE]", 1)[1]
                        if rest.strip():
                            yield f"data: {json.dumps({'token': rest})}\n\n"
                        score_sent = True
                        buf = ""
                else:
                    yield f"data: {json.dumps({'token': token})}\n\n"

        if not score_sent:
            yield f"data: {json.dumps({'score': 50})}\n\n"

        yield "data: [DONE]\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.post("/api/analyze-stream")
async def analyze_stream(body: AnalyzeStreamRequest, x_license_key: Optional[str] = Header(None)):
    """Stream iterative question-by-question analysis with prompt caching on the CV block."""
    if not body.cvText or not body.jobText:
        async def _empty():
            yield 'data: {"error": "לא נמצאו נתונים לניתוח."}\n\n'
            yield "data: [DONE]\n\n"
        return StreamingResponse(_empty(), media_type="text/event-stream",
                                 headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    license_key = x_license_key or ""
    try:
        await verify_gumroad_license(license_key)
    except Exception:
        async def _auth_err():
            yield 'data: {"error": "רישיון לא תקף."}\n\n'
            yield "data: [DONE]\n\n"
        return StreamingResponse(_auth_err(), media_type="text/event-stream",
                                 headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

    he_chars = sum(1 for c in body.jobText[:300] if "א" <= c <= "ת")
    lang = "Hebrew" if he_chars > 15 else "English"

    answers_block = ""
    if body.answers:
        lines = "\n".join(
            f"- {a.get('skill','')}: {a.get('answer','')}"
            for a in body.answers if a.get("answer") and a["answer"] != "לא ענה"
        )
        if lines:
            answers_block = f"\nCandidate self-assessment:\n{lines}\n"

    constraints_block = (
        f"\n\nUser personal constraints:\n{body.userConstraints.strip()}"
        if body.userConstraints and body.userConstraints.strip() else ""
    )

    cv_text  = body.cvText[:3000]
    job_text = body.jobText[:2500]
    sys_blocks = _cv_system_blocks(cv_text)

    # Determine gap_pct from preflight cache data embedded in answers (weight sum), fallback 20
    gap_pct  = min(40, max(10, sum(a.get("weight", 0) for a in (body.answers or [])) or 20))
    base_score_guess = 65  # used only for question prompt context
    n_q = 4 if gap_pct >= 20 else 3

    resolved_model = _resolve_model(body.model)

    question_prompt = STREAM_QUESTIONS_PROMPT.format(
        base_score=base_score_guess,
        gap_pct=gap_pct,
        n_questions=n_q,
        answers_block=answers_block,
        job_text=job_text,
    )

    analysis_prompt = STREAM_ANALYSIS_PROMPT.format(
        job_text=job_text,
        answers_block=answers_block,
        constraints_block=constraints_block,
        lang=lang,
    )

    async def generate():
        # ── Phase 1: stream questions one-by-one ──────────────────────────
        try:
            buffer = ""
            async with anthropic_client.messages.stream(
                model=resolved_model,
                max_tokens=600,
                system=sys_blocks,
                messages=[{"role": "user", "content": question_prompt}],
            ) as stream:
                async for chunk in stream.text_stream:
                    buffer += chunk
                    # Emit each complete JSON line immediately as it arrives
                    while "\n" in buffer:
                        line, buffer = buffer.split("\n", 1)
                        line = line.strip()
                        if not line:
                            continue
                        if line == "[QUESTIONS_DONE]":
                            break
                        if line.startswith("{"):
                            try:
                                q = json.loads(repair_json(line))
                                yield f"data: {json.dumps({'question': q})}\n\n"
                            except Exception:
                                pass
        except Exception as e:
            print(f"[JMA:stream_questions] {e}")

        yield f"data: {json.dumps({'phase': 'analysis'})}\n\n"

        # ── Phase 2: stream narrative analysis ────────────────────────────
        try:
            async with anthropic_client.messages.stream(
                model=resolved_model,
                max_tokens=500,
                system=sys_blocks,
                messages=[{"role": "user", "content": analysis_prompt}],
            ) as stream:
                async for chunk in stream.text_stream:
                    yield f"data: {json.dumps({'text': chunk})}\n\n"
        except Exception as e:
            print(f"[JMA:stream_analysis] {e}")
            yield f'data: {json.dumps({"text": "שגיאה בניתוח. אנא נסה שוב."})}\n\n'

        yield "data: [DONE]\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


class ScoreAnswerRequest(BaseModel):
    question: str = ""
    skill: str = ""
    answer: str = ""


@app.post("/api/score-answer")
async def score_answer(body: ScoreAnswerRequest, x_license_key: Optional[str] = Header(None)):
    """Fast single-answer scorer used for real-time FAB score updates."""
    if not body.answer.strip():
        return {"score_pct": 0}
    prompt = (
        f"Rate how well this answer demonstrates the required competency.\n\n"
        f"Skill: {body.skill}\nQuestion: {body.question}\nAnswer: {body.answer[:400]}\n\n"
        f"Score 0-100 only (0-15 none, 16-40 basic, 41-65 partial, 66-85 solid, 86-100 expert).\n"
        f'Reply JSON only: {{"score":<integer>}}'
    )
    try:
        resp = await anthropic_client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=12,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        data = _safe_json_loads(raw)
        score = max(0, min(100, int(data.get("score", 50))))
        return {"score_pct": score}
    except Exception:
        return {"score_pct": 50}


@app.post("/api/verify-license")
async def verify_license(body: VerifyLicenseRequest):
    info = await verify_gumroad_license(body.licenseKey)
    usage_count = get_usage_count(body.licenseKey)
    tier = info.get("tier", "standard")
    return {
        "success":            True,
        "email":              info["email"],
        "tier":               tier,
        "isPremium":          info.get("isPremium", False),
        "subscriptionActive": info.get("subscriptionActive", True),
        "usageCount":         usage_count,
        "monthlyLimit":       MONTHLY_USAGE_LIMIT,
        # upgradeUrl is only returned for Standard users — Premium users don't need it
        "upgradeUrl":         UPGRADE_URL if tier == "standard" else None,
    }


@app.post("/api/analyze")
async def analyze(body: AnalyzeRequest, x_license_key: Optional[str] = Header(None)):
    license_key = x_license_key or body.licenseKey or ""
    print(f"[JMA:analyze] cv_len={len(body.cvText)} job_len={len(body.jobText)} preflight={body.preflight} answers={len(body.answers)}")

    if body.preflight:
        # Two-pass preflight with CV cached in system prompt:
        # Pass 1 (fast): base score + gap_pct + full analysis fields
        # Pass 2 (fast): weighted questions using gap_pct from pass 1
        await verify_gumroad_license(license_key)

        cv_text  = body.cvText[:3000]   # cache up to 3 k chars of CV
        job_text = body.jobText[:2500]
        sys_blocks = _cv_system_blocks(cv_text)

        resolved_model = _resolve_model(body.model)
        # ── Pass 1: base analysis ─────────────────────────────────────────
        try:
            raw1, stop1 = await call_claude_cached(
                system_blocks=sys_blocks,
                user_content=BASE_ANALYSIS_USER.format(
                    scoring_rules=_SCORING_RULES,
                    job_text=job_text,
                ),
                max_tokens=1800,
                model=resolved_model,
            )
            if stop1 == "max_tokens":
                raise HTTPException(status_code=422, detail="Analysis body too long")
            analysis = parse_json_response(raw1)
            base_score: int = max(0, min(100, int(analysis.get("base_score", 50))))
            gap_pct:    int = max(0, min(40,  int(analysis.get("gap_pct",    20))))
            reasoning   = analysis.get("scoring_reasoning", "")
            print(f"[JMA:preflight] pass1 base={base_score} gap={gap_pct} model={resolved_model}")
            if reasoning:
                print(f"[JMA:scoring_reasoning]\n{reasoning}")
        except HTTPException:
            raise
        except Exception as e:
            print(f"[JMA:preflight] pass1 error: {type(e).__name__}: {e}")
            analysis   = {}
            base_score = 50
            gap_pct    = 20

        # ── Pass 2: weighted questions ────────────────────────────────────
        questions: list = []
        if gap_pct > 0:
            n_q = 4 if gap_pct >= 20 else 3
            try:
                raw2, stop2 = await call_claude_cached(
                    system_blocks=sys_blocks,
                    user_content=QUESTIONS_USER.format(
                        base_score=base_score,
                        gap_pct=gap_pct,
                        n_questions=n_q,
                        job_text=job_text,
                    ),
                    max_tokens=1200,
                    model=resolved_model,
                )
                if stop2 == "max_tokens":
                    raise HTTPException(status_code=422, detail="Analysis body too long")
                q_data   = parse_json_response(raw2)
                questions = q_data.get("questions", [])
                # Normalise weights so they sum exactly to gap_pct
                total_w = sum(q.get("weight", 0) for q in questions)
                if total_w > 0 and total_w != gap_pct:
                    scale = gap_pct / total_w
                    rem   = gap_pct
                    for i, q in enumerate(questions):
                        if i < len(questions) - 1:
                            q["weight"] = max(1, round(q.get("weight", 0) * scale))
                            rem -= q["weight"]
                        else:
                            q["weight"] = max(1, rem)
                print(f"[JMA:preflight] pass2 questions={len(questions)} weights={[q.get('weight') for q in questions]}")
            except HTTPException:
                raise
            except Exception as e:
                print(f"[JMA:preflight] pass2 error: {type(e).__name__}: {e}")

        result = {
            **analysis,
            "base_score": base_score,
            "gap_pct":    gap_pct,
            "score":      base_score,   # shown on main screen before answers
            "questions":  questions,
        }
        return {"result": result, "preflight": True}

    await require_license(license_key)

    answers_text = "\n".join(
        f"- {a.get('skill', '')}: {a.get('answer', '')}"
        for a in (body.answers or [])
        if a.get('answer') and a.get('answer') not in ('לא ענה', '')
    )
    answers_section = (
        f"=== CANDIDATE'S ANSWERS TO SCREENING QUESTIONS ===\n{answers_text}\n"
        if answers_text else ""
    )

    prompt = ANALYZE_PROMPT.format(cv_text=body.cvText, job_text=body.jobText, answers_section=answers_section)

    last_error: Exception | None = None
    for attempt in range(3):
        try:
            print(f"[JMA:analyze] attempt {attempt+1}/3")
            raw = await call_claude(prompt, max_tokens=1200)
            result = parse_json_response(raw)
            increment_usage(license_key)
            print(f"[JMA:analyze] SUCCESS score={result.get('score','?') if isinstance(result,dict) else '?'}")
            return {"result": result}
        except json.JSONDecodeError as e:
            print(f"[JMA:analyze] attempt {attempt+1} JSON error: {e}")
            last_error = e
        except Exception as e:
            print(f"[JMA:analyze] attempt {attempt+1} error: {type(e).__name__}: {e}")
            last_error = e
            if "401" in str(e) or "429" in str(e):
                break

    print(f"[JMA:analyze] FAILED after retries: {last_error}")
    raise HTTPException(status_code=500, detail=str(last_error) or "Analysis failed.")


@app.post("/api/generate-cv")
async def generate_cv(body: GenerateCVRequest, x_license_key: Optional[str] = Header(None)):
    license_key = x_license_key or body.licenseKey or ""
    await require_license(license_key)

    is_hebrew = body.jobLanguage == "hebrew"
    language = "Hebrew" if is_hebrew else "English"
    language_rule = _CV_LANG_RULE_HE if is_hebrew else _CV_LANG_RULE_EN
    language_check = _CV_LANG_CHECK_HE if is_hebrew else _CV_LANG_CHECK_EN
    answers_text = (
        "\n".join(f"{a.get('skill', '')}: {a.get('answer', '')}" for a in body.answers)
        if body.answers
        else "No additional information provided."
    )

    # Build explicit URL instruction — ensures Claude puts hyperlink URLs in [CONTACT]
    # even if it doesn't parse the raw URL embedded in the CV text.
    url_instruction = ""
    if body.cvUrls:
        url_list = "\n".join(f"  - {u}" for u in body.cvUrls)
        url_instruction = (
            f"\n\nMANDATORY CONTACT URLS — the original CV contained these hyperlinks."
            f" Every one of them MUST appear verbatim as its own line in [CONTACT]:\n{url_list}"
        )

    constraints_block = ""
    if body.userConstraints and body.userConstraints.strip():
        constraints_block = (
            f"\n\n<user_constraints>\n"
            f"These are the user's personal hard rules. They MUST be followed without exception, "
            f"even if they conflict with general CV best practices or the job description:\n"
            f"{body.userConstraints.strip()}\n"
            f"</user_constraints>"
        )

    pass1_prompt = CV_PASS1_PROMPT.format(
        language=language,
        language_rule=language_rule,
        cv_text=body.cvText,
        answers_text=answers_text,
        job_text=body.jobText,
    ) + url_instruction + constraints_block
    cv_draft = await call_claude(pass1_prompt, max_tokens=4000, model="claude-haiku-4-5-20251001", check_truncation=True)

    pass2_prompt = CV_PASS2_PROMPT.format(
        language_check=language_check,
        cv_draft=cv_draft,
        job_text=body.jobText,
        original_cv=body.cvText,
    ) + url_instruction + constraints_block
    cv_final = await call_claude(pass2_prompt, max_tokens=3800,
                                  model=_resolve_model(body.model), check_truncation=True)

    validation_issues = _validate_cv_output(cv_final, body.cvText)
    if validation_issues:
        print(f"[JMA:generate_cv] VALIDATION ISSUES: {validation_issues} — retrying Pass2 once")
        retry_prompt = pass2_prompt + (
            "\n\nIMPORTANT: your previous output had structural problems: "
            f"{', '.join(validation_issues)}. Ensure every section marker is present with "
            "real content, and do not shrink the CV drastically compared to the original."
        )
        try:
            cv_final_retry = await call_claude(retry_prompt, max_tokens=3800,
                                                 model=_resolve_model(body.model), check_truncation=True)
            if not _validate_cv_output(cv_final_retry, body.cvText):
                cv_final = cv_final_retry
                print("[JMA:generate_cv] retry succeeded")
            else:
                print("[JMA:generate_cv] retry still has issues — using original output")
        except Exception as e:
            print(f"[JMA:generate_cv] retry failed: {e} — using original output")

    # Inject tracking links only when original CV had GitHub/LinkedIn URLs and tracking is enabled
    app_id = str(uuid.uuid4())[:8]
    if body.enableTracking:
        cv_final = inject_tracking_links(cv_final, app_id)
        # Register for WebSocket push-notifications
        _app_id_registry[app_id] = {
            "user_id": _ws_user_id(license_key),
            "job_title": body.jobTitle or "",
            "company": body.company or "",
        }
    else:
        app_id = ""  # no tracking — return empty appId so client skips analytics

    # ── Diff pass: compare original CV to adapted CV, generate Hebrew explanations ──
    sections: list = []
    try:
        adapted_secs = parse_cv_sections_py(cv_final)
        sections_block_parts = []
        for idx, marker in enumerate(DIFF_SECTIONS, start=1):
            content = adapted_secs.get(marker, '').strip()
            if not content:
                continue
            label = SECTION_LABELS[marker]
            # Truncate very long sections to keep prompt size reasonable
            truncated = content[:1800] + ('…' if len(content) > 1800 else '')
            sections_block_parts.append(
                f"Section {idx} — {marker} (label: {label}, id: {idx}):\n{truncated}"
            )
        if sections_block_parts:
            diff_prompt = CV_DIFF_PROMPT.format(
                sections_block='\n\n'.join(sections_block_parts),
                original_cv=body.cvText[:7000],
                job_context=body.jobText[:200],
            )
            raw_diff = await call_claude(diff_prompt, max_tokens=2500)
            sections = parse_json_array(raw_diff)
            print(f"[JMA:diff] parsed {len(sections)} sections")
    except Exception as e:
        print(f"[JMA:diff] diff pass failed — {type(e).__name__}: {e}")
        sections = []

    # ── Cover letter pass (optional) ──────────────────────────────────────────
    cover_letter_text = ""
    if body.generateCoverLetter:
        try:
            cl_language = "Hebrew" if body.jobLanguage == "hebrew" else "English"
            cl_prompt = COVER_LETTER_PROMPT.format(
                language=cl_language,
                cv_text=body.cvText[:2000],
                job_text=body.jobText[:1500],
            )
            cover_letter_text = await call_claude(cl_prompt, max_tokens=600)
            cover_letter_text = cover_letter_text.strip()
            print(f"[JMA:cover_letter] generated {len(cover_letter_text)} chars")
        except Exception as e:
            print(f"[JMA:cover_letter] failed — {type(e).__name__}: {e}")
            cover_letter_text = ""

    increment_usage(license_key)
    return {"cvText": cv_final, "appId": app_id, "sections": sections, "coverLetterText": cover_letter_text}


@app.websocket("/ws/clicks")
async def ws_clicks(websocket: WebSocket, user_id: str = ""):
    """Persistent push channel — client receives click events without polling."""
    origin = websocket.headers.get("origin", "unknown")
    # Accept unconditionally — extension origins (chrome-extension://) must not be blocked.
    await websocket.accept()
    uid = user_id.strip()
    if uid:
        _ws_connections.setdefault(uid, set()).add(websocket)
        print(f"[JMA:ws] connected user_id={uid} origin={origin} total_users={len(_ws_connections)}")
    else:
        print(f"[JMA:ws] anonymous connection from origin={origin} (no user_id)")
    try:
        while True:
            text = await websocket.receive_text()
            try:
                msg = json.loads(text)
                if msg.get("type") == "ping":
                    await websocket.send_text(json.dumps({"type": "pong"}))
            except json.JSONDecodeError:
                pass
    except WebSocketDisconnect:
        pass
    finally:
        if uid:
            conns = _ws_connections.get(uid, set())
            conns.discard(websocket)
            if not conns:
                _ws_connections.pop(uid, None)
            print(f"[JMA:ws] disconnected user_id={uid}")


@app.get("/api/v1/clicks")
async def get_clicks(app_ids: str = ""):
    """Return click events for the given comma-separated app_ids."""
    ids = {i.strip() for i in app_ids.split(",") if i.strip()}
    try:
        all_clicks: list = json.loads(CLICKS_FILE.read_text()) if CLICKS_FILE.exists() else []
    except (json.JSONDecodeError, OSError):
        all_clicks = []
    result: dict = {}
    for c in all_clicks:
        aid = c.get("app_id", "")
        if aid in ids:
            result.setdefault(aid, []).append({
                "target": c.get("target"),
                "url": c.get("url"),
                "ts": c.get("ts"),
            })
    return {"clicks": result}


@app.get("/api/analytics/market-compare")
async def market_compare(years_exp: int = 3, title: str = "Software Engineer"):
    """
    Return anonymised market-benchmark data for the requesting developer's profile.
    Currently returns a stable mock; will be backed by aggregated opt-in telemetry.
    """
    # Tier buckets so the mock feels context-sensitive
    if years_exp <= 1:
        percentile, avg_days = 45, 14
        companies = [
            {"name": "WalkMe", "openings": 5},
            {"name": "Fiverr", "openings": 4},
            {"name": "IronSource", "openings": 3},
        ]
    elif years_exp <= 3:
        percentile, avg_days = 63, 10
        companies = [
            {"name": "Monday.com", "openings": 14},
            {"name": "Wix", "openings": 11},
            {"name": "Amdocs", "openings": 9},
        ]
    elif years_exp <= 6:
        percentile, avg_days = 79, 6
        companies = [
            {"name": "Microsoft IL", "openings": 31},
            {"name": "Google IL", "openings": 26},
            {"name": "Checkpoint", "openings": 22},
        ]
    else:
        percentile, avg_days = 93, 3
        companies = [
            {"name": "Amazon IL", "openings": 48},
            {"name": "Meta IL", "openings": 41},
            {"name": "Nvidia IL", "openings": 37},
        ]

    return {
        "percentile": percentile,
        "avg_response_days": avg_days,
        "top_trending_companies": companies,
        "week_activity": {
            "applications_avg_peers": 4,
        },
        "response_rate": {
            "market_avg_pct": 21,
        },
        "profile": {"years_exp": years_exp, "title": title},
    }


@app.get("/api/v1/track")
async def track_click(app_id: str, target: str, url: str):
    """Log a link click, push via WebSocket, and redirect to the original URL."""
    ts = datetime.utcnow().isoformat()
    try:
        clicks: list = json.loads(CLICKS_FILE.read_text()) if CLICKS_FILE.exists() else []
    except (json.JSONDecodeError, OSError):
        clicks = []
    clicks.append({"app_id": app_id, "target": target, "url": url, "ts": ts})
    try:
        CLICKS_FILE.write_text(json.dumps(clicks, indent=2))
    except OSError:
        pass

    # Push real-time notification to the owner's open WebSocket (if any)
    info = _app_id_registry.get(app_id)
    if info:
        asyncio.create_task(_ws_push(info["user_id"], {
            "type": "click",
            "app_id": app_id,
            "target": target,
            "jobTitle": info.get("job_title", ""),
            "company": info.get("company", ""),
            "ts": ts,
        }))

    return RedirectResponse(url=url, status_code=302)


@app.post("/api/rank-jobs")
async def rank_jobs(body: RankJobsRequest, x_license_key: Optional[str] = Header(None)):
    license_key = x_license_key or body.licenseKey or ""
    print(f"[JMA:rank] jobs={len(body.jobs)} cv_len={len(body.cvText)}")
    await require_license(license_key)

    def job_text(j, i):
        detail = j.get('fullText') or j.get('snippet') or ''
        company = f" at {j.get('company')}" if j.get('company') else ''
        return f"{j.get('index', i)}. {j.get('title', '?')}{company}: {detail[:1500]}"

    jobs_list = "\n\n".join(job_text(j, i) for i, j in enumerate(body.jobs[:12]))
    cv_summary = body.cvText[:800]
    prompt = RANK_JOBS_PROMPT.format(cv_summary=cv_summary, jobs_list=jobs_list)
    print(f"[JMA:rank] prompt_len={len(prompt)} jobs_included={min(len(body.jobs),12)}")

    last_error: Exception | None = None
    for attempt in range(2):
        try:
            print(f"[JMA:rank] attempt {attempt+1}/2")
            raw = await call_claude(prompt, max_tokens=1200)
            ranked = parse_json_response(raw)
            if not isinstance(ranked, list):
                print(f"[JMA:rank] response is not a list: {type(ranked)} val={str(ranked)[:100]}")
                raise ValueError(f"Expected JSON array, got {type(ranked).__name__}: {str(ranked)[:80]}")

            jobs_map = {j.get('index', i): j for i, j in enumerate(body.jobs)}
            result = []
            for item in ranked:
                idx = item.get('index', 0)
                original = jobs_map.get(idx) or (body.jobs[idx] if idx < len(body.jobs) else {})
                result.append({
                    "index": idx,
                    "title": original.get('title', ''),
                    "company": original.get('company', ''),
                    "score": max(0, min(100, int(item.get('score', 50)))),
                    "pro": item.get('pro', ''),
                    "con": item.get('con', ''),
                })
            increment_usage(license_key)
            print(f"[JMA:rank] SUCCESS returned {len(result)} ranked jobs")
            return {"rankedJobs": result}
        except Exception as e:
            print(f"[JMA:rank] attempt {attempt+1} error: {type(e).__name__}: {e}")
            last_error = e

    print(f"[JMA:rank] FAILED: {last_error}")
    raise HTTPException(status_code=500, detail=str(last_error) or "Ranking failed.")


@app.post("/api/scrape-job")
async def scrape_job(body: ScrapeJobRequest):
    """Crowdsourced silent job scraping — no auth, no AI. Just dedup + store."""
    url = body.url.strip()
    if not url or not url.startswith("http"):
        raise HTTPException(status_code=400, detail="Invalid URL")

    jobs = _load_raw_jobs()
    existing_urls = {j["url"] for j in jobs}
    if url in existing_urls:
        return {"status": "duplicate"}

    jobs.append({
        "url": url,
        "text": body.text[:5000],
        "title": (body.title or "")[:200],
        "ts": datetime.utcnow().isoformat(),
    })

    if len(jobs) > MAX_RAW_JOBS:
        jobs = jobs[-MAX_RAW_JOBS:]

    _save_raw_jobs(jobs)
    return {"status": "saved"}


@app.post("/api/import-jobs")
async def import_jobs(body: ImportJobsRequest, x_license_key: Optional[str] = Header(None)):
    """Premium: 2-stage filtering of scraped jobs → Excel download."""
    license_key = x_license_key or ""

    # Real-time Gumroad premium check (TEST_LICENSE_KEY bypasses instantly)
    if not await _verify_premium(license_key):
        raise HTTPException(status_code=403, detail="דרוש רישיון פרימיום לפיצ'ר זה.")

    all_jobs = _load_raw_jobs()
    if not all_jobs:
        raise HTTPException(
            status_code=404,
            detail="עדיין לא נאספו משרות. המשיכי לגלוש — כל משרה שתבקרי בה תיאסף אוטומטית.",
        )

    # Time filter
    if body.timeRange == "since_last":
        usage = _load_usage()
        cutoff = usage.get(license_key, {}).get("last_import_ts", "")
        if not cutoff:
            cutoff = (datetime.utcnow() - timedelta(days=30)).isoformat()
    else:
        cutoff = (datetime.utcnow() - timedelta(days=3)).isoformat()

    time_filtered = [j for j in all_jobs if (j.get("ts") or "") >= cutoff]
    if not time_filtered:
        raise HTTPException(status_code=404, detail="לא נמצאו משרות חדשות בטווח הזמן שנבחר.")

    # Step 1: fast keyword filter — no AI, no cost
    step1 = _quick_filter(time_filtered, body.cvText)
    if not step1:
        raise HTTPException(status_code=404, detail="לא נמצאו משרות רלוונטיות לאחר סינון ראשוני.")

    step1 = step1[:100]  # cost cap before Claude

    # Step 2: concurrent per-job Claude scoring (semaphore limits to 5 parallel calls)
    cv_summary = body.cvText[:800]
    sem = asyncio.Semaphore(5)

    async def rank_with_sem(job: dict):
        async with sem:
            return await _rank_single_job(cv_summary, job)

    results = await asyncio.gather(*[rank_with_sem(j) for j in step1], return_exceptions=True)

    ranked: list[dict] = []
    for result in results:
        if not isinstance(result, dict):
            continue
        if result["score"] >= body.minScore:
            job = result["job"]
            ranked.append({
                "Title": job.get("title", ""),
                "URL": job.get("url", ""),
                "Score": result["score"],
                "Pro": result["pro"],
                "Con": result["con"],
                "Date": (job.get("ts") or "")[:10],
            })

    if not ranked:
        raise HTTPException(
            status_code=404,
            detail=f"לא נמצאו משרות שעוברות את אחוז ההתאמה המינימלי ({body.minScore}%). נסי להוריד את הסף.",
        )

    ranked.sort(key=lambda x: x["Score"], reverse=True)

    # Save last import timestamp
    usage = _load_usage()
    usage.setdefault(license_key, {})["last_import_ts"] = datetime.utcnow().isoformat()
    _save_usage(usage)

    # Generate Excel
    try:
        import pandas as pd
        from openpyxl.styles import Alignment, Font, PatternFill

        df = pd.DataFrame(ranked, columns=["Title", "URL", "Score", "Pro", "Con", "Date"])
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Matched Jobs")
            ws = writer.sheets["Matched Jobs"]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor="7C3AED")
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        output.seek(0)
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel generation unavailable on this server.")

    filename = f"JobMatchAI_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", "8000"))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
